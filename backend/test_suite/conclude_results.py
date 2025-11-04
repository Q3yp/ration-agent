#!/usr/bin/env python3
"""
Summarise a batch test run into a single text file.
For each successful scenario, capture:
  • the sanitized prompt delivered to the agent
  • the default feedbase details returned by the tooling
  • the exported formulation summary from the Excel file
The output is written alongside the run artefacts as results_summary.txt.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook


TEST_SUITE_DIR = Path(__file__).parent
TEST_RUNS_DIR = TEST_SUITE_DIR / "test_runs"


def _load_manifest(run_dir: Path) -> dict:
    manifest_path = run_dir / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"manifest.json not found in {run_dir}")
    with manifest_path.open(encoding="utf-8") as fh:
        return json.load(fh)


def _find_unique_file(directory: Path, pattern: str) -> Optional[Path]:
    matches = list(directory.glob(pattern))
    if not matches:
        return None
    if len(matches) > 1:
        # Prefer most recently modified if duplicates somehow exist.
        matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0]


def _load_prompt(metadata_path: Path, messages_path: Optional[Path]) -> str:
    prompt = ""
    if metadata_path and metadata_path.exists():
        with metadata_path.open(encoding="utf-8") as fh:
            metadata = json.load(fh)
        prompt = metadata.get("prompt", "").strip()
    if prompt or not messages_path:
        return prompt

    with messages_path.open(encoding="utf-8") as fh:
        transcript = json.load(fh)
    for message in transcript.get("messages", []):
        if message.get("type") == "human":
            content = message.get("content", "")
            if content:
                return content.strip()
    return prompt


def _load_feedbase_text(messages_path: Optional[Path]) -> str:
    if not messages_path or not messages_path.exists():
        return "未找到 feedbase 信息。"

    with messages_path.open(encoding="utf-8") as fh:
        transcript = json.load(fh)

    def _first_matching(names: Iterable[str]) -> Optional[str]:
        for msg in transcript.get("messages", []):
            if msg.get("name") in names and msg.get("content"):
                return msg["content"].strip()
        return None

    feedbase_text = _first_matching({"check_feeds"})
    if not feedbase_text:
        feedbase_text = _first_matching({"list_feed_bases"})
    return feedbase_text or "未找到 feedbase 信息。"


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        formatted = f"{value:.3f}".rstrip("0").rstrip(".")
        return formatted if formatted else "0"
    return str(value)


def _format_markdown_table(headers: List[str], rows: List[List[str]]) -> List[str]:
    if not headers or not rows:
        return []
    header_line = "| " + " | ".join(headers) + " |"
    divider = "| " + " | ".join(["---"] * len(headers)) + " |"
    table_lines = [header_line, divider]
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        table_lines.append("| " + " | ".join(padded[:len(headers)]) + " |")
    return table_lines


def _extract_formulation_details(excel_path: Optional[Path]) -> dict:
    if not excel_path or not excel_path.exists():
        return {}

    workbook = load_workbook(excel_path, data_only=True)
    details: dict = {
        "summary": "",
        "composition_headers": [],
        "composition_rows": [],
        "nutrition_headers": [],
        "nutrition_rows": []
    }

    if "配方说明" in workbook.sheetnames:
        ws = workbook["配方说明"]
        summary_lines: List[str] = []
        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            summary_lines.append("" if cell is None else str(cell).rstrip())
        while summary_lines and not summary_lines[0]:
            summary_lines.pop(0)
        while summary_lines and not summary_lines[-1]:
            summary_lines.pop()
        details["summary"] = "\n".join(summary_lines).strip()

    sheet_name = "配方结果" if "配方结果" in workbook.sheetnames else workbook.sheetnames[0]
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    composition_records: List[dict] = []
    idx = 0
    while idx < len(rows):
        row = rows[idx]
        first = ""
        if row and row[0] is not None:
            first = str(row[0]).strip()

        if first == "饲料名称":
            header_primary = [str(v).strip() if v else "" for v in row]
            header_secondary = []
            if idx + 1 < len(rows):
                header_secondary = [str(v).strip() if v else "" for v in rows[idx + 1]]
            column_names: List[str] = []
            for col_idx in range(len(header_primary)):
                primary = header_primary[col_idx]
                secondary = header_secondary[col_idx] if col_idx < len(header_secondary) else ""
                name = secondary if secondary else primary
                if not name:
                    name = f"列{col_idx + 1}"
                column_names.append(name)
            idx += 2

            while idx < len(rows):
                data_row = rows[idx]
                if not data_row or all(value is None for value in data_row):
                    idx += 1
                    continue
                first_cell = str(data_row[0]).strip() if data_row[0] is not None else ""
                if first_cell in ("整体营养成分分析", ""):
                    break
                record = {}
                for col_idx, name in enumerate(column_names):
                    value = data_row[col_idx] if col_idx < len(data_row) else ""
                    record[name] = _format_cell(value)
                composition_records.append(record)
                idx += 1
            continue

        if first == "整体营养成分分析":
            idx += 1
            if idx >= len(rows):
                break
            header_row = rows[idx]
            nutrition_headers = [str(v).strip() for v in header_row if v not in (None, "")]
            if not nutrition_headers:
                nutrition_headers = ["营养成分", "含量"]
            idx += 1
            nutrition_rows: List[List[str]] = []
            while idx < len(rows):
                data_row = rows[idx]
                if not data_row or all(value is None for value in data_row):
                    break
                values = [_format_cell(data_row[col_idx] if col_idx < len(data_row) else "") for col_idx in range(len(nutrition_headers))]
                nutrition_rows.append(values)
                idx += 1
            details["nutrition_headers"] = nutrition_headers
            details["nutrition_rows"] = nutrition_rows
            break

        idx += 1

    desired_columns = [
        "饲料名称",
        "日饲喂量 (kg/day)",
        "干物质比例 (%)",
        "CP",
        "Ca",
        "P",
        "NEm_Mcal",
        "NEg_Mcal"
    ]

    if composition_records:
        available_columns = [column for column in desired_columns if column in composition_records[0]]
        if not available_columns:
            available_columns = list(composition_records[0].keys())
        composition_rows = [
            [record.get(column, "") for column in available_columns]
            for record in composition_records
        ]
        details["composition_headers"] = available_columns
        details["composition_rows"] = composition_rows

    if "summary" not in details or not details["summary"]:
        details["summary"] = "未找到配方说明。"

    return details


def _scenario_directory(results_dir: Path, scenario_id: int, scenario_name: str) -> Optional[Path]:
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", scenario_name)
    candidate = results_dir / "success" / f"场景{scenario_id:02d}_{safe_name}"
    if candidate.exists():
        return candidate

    # Fallback: match by prefix if exact unicode differs (e.g., trimmed whitespace).
    success_dir = results_dir / "success"
    if not success_dir.exists():
        return None
    for path in success_dir.iterdir():
        if path.is_dir() and path.name.startswith(f"场景{scenario_id:02d}"):
            return path
    return None


def summarise_run(timestamp: str, output_name: str = "results_summary.txt") -> Path:
    run_dir = TEST_RUNS_DIR / timestamp
    if not run_dir.exists():
        raise FileNotFoundError(f"测试运行目录不存在: {run_dir}")

    manifest = _load_manifest(run_dir)
    results_dir = run_dir / "results"
    summary_lines: List[str] = []
    scenario_entries = manifest.get("scenarios", [])

    meta = manifest.get("metadata", {})
    summary_lines.append("=" * 80)
    summary_lines.append(f"测试批次: {timestamp}")
    summary_lines.append(f"动物类型: {meta.get('animal_type', 'unknown')}")
    total_count = meta.get("total_scenarios", len(scenario_entries))
    success_count = meta.get("successful", sum(1 for s in scenario_entries if s.get("status") == "success"))
    failure_count = meta.get("failed", total_count - success_count)

    summary_lines.append(f"总场景数: {total_count}")
    summary_lines.append(f"成功: {success_count}  失败: {failure_count}")
    summary_lines.append(f"测试账号: {meta.get('test_account', 'unknown')}")
    scenario_sections: List[List[str]] = []
    feedbase_reference: Optional[str] = None
    feedbase_name: Optional[str] = None

    for scenario in scenario_entries:
        scenario_id = scenario.get("scenario_id")
        scenario_name = scenario.get("scenario_name", "未知场景")
        status = scenario.get("status", "unknown")
        session_id = scenario.get("session_id", "unknown")

        status_display = {
            "success": "成功",
            "submitted": "已提交",
            "failed": "失败"
        }.get(status, status)

        section_lines: List[str] = []
        section_lines.append(f"场景 {scenario_id:02d}: {scenario_name}")
        section_lines.append(f"状态: {status_display}")
        section_lines.append(f"Session ID: {session_id}")

        if status != "success":
            safe_name = re.sub(r'[\\/:*?"<>|]', "_", scenario_name)
            failure_dir = results_dir / "failed" / f"场景{scenario_id:02d}_{safe_name}"
            metadata_path = failure_dir / "metadata.json"
            failure_reason = None
            prompt_text = scenario.get("prompt", "").strip()
            if metadata_path.exists():
                with metadata_path.open(encoding="utf-8") as fh:
                    metadata = json.load(fh)
                failure_reason = metadata.get("failure_reason", failure_reason)
                prompt_text = metadata.get("prompt", prompt_text).strip()
            if prompt_text:
                section_lines.append("")
                section_lines.append("【配方任务提示】")
                section_lines.extend(prompt_text.splitlines())

            section_lines.append("")
            section_lines.append("【结果说明】")
            section_lines.append(failure_reason or "未生成有效配方。")
            section_lines.append("-" * 80)
            scenario_sections.append(section_lines)
            continue

        scenario_dir = _scenario_directory(results_dir, scenario_id, scenario_name)
        if not scenario_dir:
            section_lines.append("未找到成功结果目录。")
            section_lines.append("-" * 80)
            scenario_sections.append(section_lines)
            continue

        metadata_path = scenario_dir / "metadata.json"
        messages_path = _find_unique_file(scenario_dir, "*_messages.json")
        excel_path = _find_unique_file(scenario_dir, "*.xlsx")

        prompt_text = _load_prompt(metadata_path, messages_path)
        feedbase_text = _load_feedbase_text(messages_path) if messages_path else None

        if feedbase_text and "未找到 feedbase 信息" not in feedbase_text and not feedbase_reference:
            feedbase_reference = feedbase_text
            match = re.search(r"Feedbase '([^']+)'", feedbase_text)
            if match:
                feedbase_name = match.group(1)

        formulation_details = _extract_formulation_details(excel_path)
        summary_text = formulation_details.get("summary", "")
        composition_headers = formulation_details.get("composition_headers", [])
        composition_rows = formulation_details.get("composition_rows", [])
        nutrition_headers = formulation_details.get("nutrition_headers", [])
        nutrition_rows = formulation_details.get("nutrition_rows", [])

        if prompt_text:
            section_lines.append("")
            section_lines.append("【配方任务提示】")
            section_lines.extend(prompt_text.splitlines())

        if feedbase_name:
            section_lines.append("")
            section_lines.append(f"默认饲料库: {feedbase_name}（详见上文）")

        if composition_headers and composition_rows:
            section_lines.append("")
            section_lines.append("【配方组成（关键指标）】")
            section_lines.extend(_format_markdown_table(composition_headers, composition_rows))

        if nutrition_headers and nutrition_rows:
            section_lines.append("")
            section_lines.append("【营养成分分析】")
            section_lines.extend(_format_markdown_table(nutrition_headers, nutrition_rows))

        if summary_text:
            section_lines.append("")
            section_lines.append("【配方说明】")
            section_lines.extend(summary_text.splitlines())

        section_lines.append("-" * 80)
        scenario_sections.append(section_lines)

    summary_lines.append("=" * 80)
    summary_lines.append("")

    if feedbase_reference:
        summary_lines.append("【默认饲料库】")
        summary_lines.append(feedbase_reference)
        summary_lines.append("")

    for section in scenario_sections:
        summary_lines.extend(section)
        summary_lines.append("")

    output_path = run_dir / output_name
    with output_path.open("w", encoding="utf-8") as fh:
        fh.write("\n".join(summary_lines))

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Summarise a test batch run into a single text file.")
    parser.add_argument("timestamp", help="Test run timestamp (e.g. 20251031_100727)")
    parser.add_argument(
        "--output",
        default="results_summary.txt",
        help="Output filename (written inside the test run directory)"
    )
    args = parser.parse_args()

    output_path = summarise_run(args.timestamp, args.output)
    print(f"✓ 汇总结果已写入: {output_path}")


if __name__ == "__main__":
    main()
