#!/usr/bin/env python3
"""
Extract test results for a given test run timestamp.
Collects exported formulation files and session data.
"""

import asyncio
import json
import shutil
import sys
from pathlib import Path
from typing import Dict, List

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from dump_session import dump_session
from conclude_results import summarise_run
from openpyxl import load_workbook
import re


def has_valid_formulation(excel_path: Path) -> bool:
    """Return True if the workbook contains a non-empty formulation."""
    try:
        workbook = load_workbook(excel_path, data_only=True)
    except Exception:
        return False

    if "配方说明" in workbook.sheetnames:
        sheet = workbook["配方说明"]
        content_lines = 0
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            if any(str(cell).strip() for cell in row if cell is not None):
                content_lines += 1
            if content_lines >= 5:
                return True

    sheet_name = "配方结果" if "配方结果" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    data_rows = 0
    for row in sheet.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if "饲料名称" in values and "日饲喂量" in "".join(values):
            continue
        if all(not value for value in values):
            continue
        data_rows += 1
        if data_rows >= 1:
            return True
    return False


async def extract_test_results(timestamp: str):
    """
    Extract all results for a test run.

    Args:
        timestamp: Test run timestamp (e.g., '20251030_120000')
    """
    # Define paths
    test_suite_dir = Path(__file__).parent
    test_runs_dir = test_suite_dir / "test_runs"
    test_run_dir = test_runs_dir / timestamp
    files_dir = Path(__file__).parent.parent / "files"

    # Check if test run exists
    if not test_run_dir.exists():
        print(f"❌ Test run not found: {test_run_dir}")
        print(f"\nAvailable test runs:")
        if test_runs_dir.exists():
            for run in sorted(test_runs_dir.iterdir(), reverse=True):
                if run.is_dir():
                    print(f"  - {run.name}")
        sys.exit(1)

    # Load manifest
    manifest_file = test_run_dir / "manifest.json"
    if not manifest_file.exists():
        print(f"❌ Manifest not found: {manifest_file}")
        sys.exit(1)

    with open(manifest_file, 'r', encoding='utf-8') as f:
        manifest = json.load(f)

    print("="*80)
    print(f"提取测试结果 - {timestamp}")
    print("="*80)
    print(f"测试账号: {manifest['metadata'].get('test_account', 'unknown')}")
    print(f"动物类型: {manifest['metadata'].get('animal_type', 'unknown')}")
    print(f"场景总数: {manifest['metadata'].get('total_scenarios', 0)}")
    print(f"成功场景: {manifest['metadata'].get('successful', 0)}")
    print(f"失败场景: {manifest['metadata'].get('failed', 0)}")
    print("="*80)

    # Create results directory structure
    results_dir = test_run_dir / "results"
    results_dir.mkdir(exist_ok=True)

    # Create subdirectories
    (results_dir / "success").mkdir(exist_ok=True)
    (results_dir / "failed").mkdir(exist_ok=True)
    (results_dir / "formulations").mkdir(exist_ok=True)

    scenario_entries = manifest.get("scenarios", [])

    successful_extractions: List[Dict] = []
    failed_extractions: List[Dict] = []
    successful_ids = set()

    print(f"\n📦 处理 {len(scenario_entries)} 个场景...\n")

    for scenario in scenario_entries:
        scenario_id = scenario['scenario_id']
        scenario_name = scenario['scenario_name']
        session_id = scenario.get('session_id')
        status = scenario.get('status', 'submitted')
        prompt_text = scenario.get("prompt", "")
        timestamp_value = scenario.get("timestamp", "")

        print(f"场景 {scenario_id:2d}: {scenario_name}")

        scenario_success = False
        failure_reason = None
        session_files_dir = files_dir / session_id if session_id else None
        excel_files: List[Path] = []
        valid_excel_files: List[Path] = []

        if not session_id:
            failure_reason = "缺少会话 ID"
        else:
            if not session_files_dir or not session_files_dir.exists():
                failure_reason = f"会话文件目录不存在: {session_files_dir}"
            else:
                excel_files = list(session_files_dir.glob("*.xlsx"))
                valid_excel_files = [excel for excel in excel_files if has_valid_formulation(excel)]
                if not valid_excel_files:
                    failure_reason = "未找到有效配方文件"
                else:
                    scenario_success = True

        target_bucket = "success" if scenario_success else "failed"
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", scenario_name)
        scenario_dir_name = f"场景{scenario_id:02d}_{safe_name}"
        scenario_dir = results_dir / target_bucket / scenario_dir_name
        scenario_dir.mkdir(exist_ok=True)

        # Dump session data for inspection (even for failures if session exists)
        if session_id:
            try:
                print("  📥 导出会话数据...")
                await dump_session(session_id, scenario_dir, include_artifacts=True)
            except Exception as dump_error:
                print(f"  ⚠️ 会话导出失败: {dump_error}")
                if scenario_success:
                    # Treat as recoverable; keep success status but note warning
                    failure_reason = f"会话导出失败: {dump_error}"
                    scenario_success = False
                    target_bucket = "failed"
                    scenario_dir = results_dir / target_bucket / scenario_dir_name
                    scenario_dir.mkdir(exist_ok=True)

        if scenario_success:
            print(f"  📊 发现 {len(valid_excel_files)} 个有效配方文件")
            for excel_file in valid_excel_files:
                dest_file = scenario_dir / excel_file.name
                shutil.copy2(excel_file, dest_file)
                print(f"    ✓ {excel_file.name}")

                formulation_file = results_dir / "formulations" / f"场景{scenario_id:02d}_{excel_file.name}"
                shutil.copy2(excel_file, formulation_file)
        else:
            if excel_files and not valid_excel_files:
                print("  ⚠️ 检测到配方文件但内容无效，将文件复制至失败目录以供检查")
                for excel_file in excel_files:
                    dest_file = scenario_dir / excel_file.name
                    shutil.copy2(excel_file, dest_file)

        scenario_metadata = {
            "scenario_id": scenario_id,
            "scenario_name": scenario_name,
            "session_id": session_id,
            "prompt": prompt_text,
            "timestamp": timestamp_value,
            "status": "success" if scenario_success else "failed"
        }
        if failure_reason:
            scenario_metadata["failure_reason"] = failure_reason
        scenario_metadata_file = scenario_dir / "metadata.json"
        with open(scenario_metadata_file, 'w', encoding='utf-8') as f:
            json.dump(scenario_metadata, f, indent=2, ensure_ascii=False)

        if scenario_success:
            successful_ids.add(scenario_id)
            successful_extractions.append({
                "scenario_id": scenario_id,
                "scenario_name": scenario_name,
                "session_id": session_id,
                "formulation_files": len(valid_excel_files)
            })
            print("  ✅ 提取完成\n")
        else:
            failure_entry = {
                "scenario_id": scenario_id,
                "scenario_name": scenario_name,
                "session_id": session_id,
                "reason": failure_reason or "未知原因"
            }
            failed_extractions.append(failure_entry)
            print(f"  ❌ 提取失败: {failure_entry['reason']}\n")

    # Generate extraction report
    report = {
        "timestamp": timestamp,
        "total_scenarios": len(scenario_entries),
        "successful_extractions": len(successful_extractions),
        "failed_extractions": len(failed_extractions),
        "success_details": successful_extractions,
        "failed_details": failed_extractions,
    }

    report_file = test_run_dir / "extraction_report.json"
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    # Print summary
    print("="*80)
    print("提取完成汇总")
    print("="*80)
    print(f"\n成功提取: {len(successful_extractions)}/{len(manifest['scenarios'])}")
    print(f"失败: {len(failed_extractions)}/{len(manifest['scenarios'])}")

    if successful_extractions:
        print(f"\n✅ 成功提取的场景:")
        for item in successful_extractions:
            print(f"  场景{item['scenario_id']:2d} ({item['scenario_name']}): {item['formulation_files']} 个配方文件")

    if failed_extractions:
        print(f"\n❌ 提取失败的场景:")
        for item in failed_extractions:
            print(f"  场景{item['scenario_id']:2d} ({item['scenario_name']}): {item['reason']}")

    # Update manifest statuses based on actual results
    for scenario in scenario_entries:
        scenario_id = scenario['scenario_id']
        scenario['status'] = "success" if scenario_id in successful_ids else "failed"

    manifest_metadata = manifest.get("metadata", {})
    manifest_metadata["successful"] = len(successful_ids)
    manifest_metadata["failed"] = len(scenario_entries) - len(successful_ids)
    manifest["metadata"] = manifest_metadata

    manifest_file = test_run_dir / "manifest.json"
    with open(manifest_file, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    # Refresh session mapping with successful scenarios only
    mapping_file = test_run_dir / "session_mapping.json"
    session_mapping = {
        scenario["scenario_name"]: scenario["session_id"]
        for scenario in scenario_entries
        if scenario["status"] == "success" and scenario.get("session_id")
    }
    with open(mapping_file, 'w', encoding='utf-8') as f:
        json.dump(session_mapping, f, indent=2, ensure_ascii=False)

    # Generate consolidated summary
    try:
        summary_path = summarise_run(timestamp)
        print(f"\n📄 汇总文件已生成: {summary_path}")
    except Exception as summary_error:
        print(f"\n⚠️ 生成汇总文件失败: {summary_error}")

    print(f"\n{'='*80}")
    print(f"结果目录: {results_dir}")
    print(f"提取报告: {report_file}")
    print(f"{'='*80}\n")

    # Run cost analysis
    print("📊 计算成本分析...\n")
    await run_cost_analysis(test_run_dir)


async def run_cost_analysis(test_run_dir: Path):
    """Run cost analysis on extracted test results"""
    from calculate_test_costs import (
        calculate_cost,
        extract_tokens_from_messages,
        HAIKU_4_5_INPUT_PRICE,
        HAIKU_4_5_OUTPUT_PRICE
    )

    results_dir = test_run_dir / "results" / "success"
    if not results_dir.exists():
        print("⚠️  No successful results to analyze")
        return

    total_input = 0
    total_output = 0
    total_tokens = 0
    scenario_costs = []

    for scenario_dir in sorted(results_dir.iterdir()):
        if not scenario_dir.is_dir():
            continue

        # Find messages.json file
        messages_files = list(scenario_dir.glob("*_messages.json"))
        if not messages_files:
            continue

        messages_file = messages_files[0]
        input_tokens, output_tokens, tokens = extract_tokens_from_messages(messages_file)
        cost = calculate_cost(input_tokens, output_tokens)

        total_input += input_tokens
        total_output += output_tokens
        total_tokens += tokens

        scenario_costs.append({
            "scenario": scenario_dir.name,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": tokens,
            "cost": cost
        })

    # Save cost analysis
    cost_analysis = {
        "pricing": {
            "input_price_per_1m": HAIKU_4_5_INPUT_PRICE,
            "output_price_per_1m": HAIKU_4_5_OUTPUT_PRICE,
            "model": "claude-haiku-4.5"
        },
        "summary": {
            "total_scenarios": len(scenario_costs),
            "total_input_tokens": total_input,
            "total_output_tokens": total_output,
            "total_tokens": total_tokens,
            "total_cost": calculate_cost(total_input, total_output),
            "average_cost_per_scenario": calculate_cost(total_input, total_output) / len(scenario_costs) if scenario_costs else 0,
            "average_tokens_per_scenario": total_tokens / len(scenario_costs) if scenario_costs else 0
        },
        "scenarios": sorted(scenario_costs, key=lambda x: x['cost'], reverse=True)
    }

    cost_file = test_run_dir / "cost_analysis.json"
    with open(cost_file, 'w', encoding='utf-8') as f:
        json.dump(cost_analysis, f, indent=2, ensure_ascii=False)

    print(f"成本分析:")
    print(f"  总场景数: {len(scenario_costs)}")
    print(f"  总输入 tokens: {total_input:,}")
    print(f"  总输出 tokens: {total_output:,}")
    print(f"  总 tokens: {total_tokens:,}")
    print(f"  总成本: ${cost_analysis['summary']['total_cost']:.4f}")
    print(f"  平均成本/场景: ${cost_analysis['summary']['average_cost_per_scenario']:.4f}")
    print(f"\n💾 成本分析已保存: {cost_file}\n")


async def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print("Usage: python extract_test_results.py <timestamp>")
        print("\nExample:")
        print("  python extract_test_results.py 20251030_120000")
        print("\nAvailable test runs:")
        test_runs_dir = Path(__file__).parent / "test_runs"
        if test_runs_dir.exists():
            for run in sorted(test_runs_dir.iterdir(), reverse=True):
                if run.is_dir():
                    print(f"  - {run.name}")
        sys.exit(1)

    timestamp = sys.argv[1]
    await extract_test_results(timestamp)


if __name__ == "__main__":
    asyncio.run(main())
