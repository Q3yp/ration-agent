"""
Formulation Export Tool

Contains the export_formulation tool for exporting formulations to Excel.
Extracted from formulation_tools.py to reduce file size.
"""
import asyncio
import json
import logging
import re
from typing import Dict, Any, Optional, Annotated
from datetime import datetime
from pathlib import Path
import pandas as pd
from langchain_core.tools import tool, InjectedToolCallId
from langchain_core.messages import ToolMessage
from langchain_core.runnables import RunnableConfig
from langgraph.prebuilt import InjectedState
from langgraph.types import Command
from langgraph.config import get_store

from utils.language import normalize_locale, get_export_texts

logger = logging.getLogger(__name__)


def sanitize_feed_name(name: str) -> str:
    """Sanitize feed names for Excel export by removing control characters."""
    if not isinstance(name, str):
        return str(name)
    sanitized = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', name)
    sanitized = sanitized.replace('\x00', '').strip()
    return sanitized if sanitized else "饲料"


def create_export_formulation_tool(animal_type: str = "dairy_cow"):
    """Factory function to create the export_formulation tool with animal_type closure."""
    
    @tool
    async def export_formulation(
        description: str,
        state: Annotated[dict, InjectedState],
        tool_call_id: Annotated[str, InjectedToolCallId],
        config: RunnableConfig,
        filename: Optional[str] = None
    ) -> Command:
        """
        Export current formulation to Excel with multi-tab layout.

        Uses stored animal_params from set_animal_params for NASEM evaluation.
        No need to re-enter animal parameters - they are retrieved from session state.

        Creates Excel with tabs:
        - Summary: Feed formula, key nutrients, constraints, profitability, NASEM summary
        - NASEM-Intakes, NASEM-Requirements, NASEM-Production, etc. (dairy_cow only)

        Args:
            description: Detailed formulation description and recommendations (supports multi-line text)
            filename: Optional custom filename (default: formulation_export_TIMESTAMP.xlsx)

        Returns:
            Excel file with feeding information
        """
        try:
            # Get preferred language and texts
            preferred_language = config.get("configurable", {}).get("preferred_language", "zh-CN")
            locale = normalize_locale(preferred_language)
            texts = get_export_texts(locale)

            # Get all required data from state
            current_formulation = state.get("current_formulation", {})
            formulation_constraints = state.get("formulation_constraints", [])
            feed_constraints = state.get("feed_constraints", {})
            animal_params = state.get("animal_params", {})

            # Get feedbase references from state
            feedbase_name = state.get("current_feedbase_name", "")
            user_id = state.get("current_user_id", "")

            if not current_formulation or current_formulation.get("status") != "success":
                return Command(
                    update={"messages": [ToolMessage(texts["error_no_formulation"], tool_call_id=tool_call_id)]}
                )

            # Get feed database from store using state references
            feed_database = {}
            if feedbase_name and user_id:
                try:
                    store = get_store()

                    # Try user feedbase first
                    namespace = ("feedbases", user_id, feedbase_name)
                    feedbase_entry = await store.aget(namespace, "data")

                    # If not found, try system feedbase
                    if not feedbase_entry and feedbase_name.startswith("default_"):
                        system_namespace = ("system_feedbases", feedbase_name)
                        feedbase_entry = await store.aget(system_namespace, "data")

                    if feedbase_entry:
                        feedbase_data = feedbase_entry.value
                        feed_database = feedbase_data.get("feeds", {})
                except Exception as e:
                    logger.error(f"Failed to get feedbase data from store: {e}")

            if not feed_database:
                return Command(
                    update={"messages": [ToolMessage(texts["error_no_db"], tool_call_id=tool_call_id)]}
                )

            # Get session workspace path from config
            try:
                session_id = config["configurable"].get("thread_id")
                if session_id:
                    from services.session_manager import session_manager
                    workspace_path = await session_manager.get_session_workspace_path(session_id)
                else:
                    logger.warning("No session_id in config, using current directory")
                    workspace_path = Path(".")
            except Exception as e:
                logger.error(f"Failed to get session workspace: {e}")
                workspace_path = Path(".")

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"formulation_export_{timestamp}.xlsx"

            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = workspace_path / filename

            # Capture daily intake information if available (check both old and new key names)
            daily_intake_kg = current_formulation.get("daily_dm_intake_kg") or current_formulation.get("predicted_dmi_kg")
            predicted_mp_g = current_formulation.get("predicted_mp_g")  # MP supply from optimizer
            predicted_me_mcal = current_formulation.get("predicted_me_mcal")  # ME supply from optimizer
            nutrient_analysis = current_formulation.get("nutrient_analysis", {})
            formulation_feeds = current_formulation.get("formulation", {})

            # Helper function to validate constraints and calculate achieved values
            def validate_constraints():
                constraint_results = []

                for constraint in formulation_constraints:
                    constraint_type = constraint.get("type", "")
                    result = {
                        "type": "",
                        "nutrient": "",
                        "condition": "",
                        "actual": "",
                        "unit": "",
                        "satisfaction": ""
                    }
                    
                    if constraint_type == "concentration":
                        nutrient = constraint["nutrient"]
                        min_val = constraint.get("min")
                        max_val = constraint.get("max")
                        achieved = nutrient_analysis.get(nutrient, 0.0)
                        
                        result["type"] = texts["con_concentration"]
                        result["nutrient"] = nutrient
                        result["actual"] = achieved
                        result["unit"] = "% DM"
                        
                        if min_val is not None and max_val is not None:
                            result["condition"] = f"{min_val} - {max_val}"
                            satisfied = min_val <= achieved <= max_val
                        elif min_val is not None:
                            result["condition"] = f"≥ {min_val}"
                            satisfied = achieved >= min_val
                        elif max_val is not None:
                            result["condition"] = f"≤ {max_val}"
                            satisfied = achieved <= max_val
                        else:
                            satisfied = True
                            
                        result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    elif constraint_type == "daily_total":
                        attribute = constraint.get("attribute")
                        target = constraint.get("target")
                        tolerance_percent = constraint.get("tolerance_percent", 10.0)
                        
                        if attribute == "dmi":
                            result["type"] = texts["con_dmi"]
                            result["nutrient"] = "DMI"
                            result["unit"] = "kg/day"
                            result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}%"
                            result["actual"] = f"{target} ({texts['range']})"
                            result["satisfaction"] = texts["satisfied"]
                        elif attribute == "mp":
                            # Metabolizable Protein constraint - use predicted_mp_g from optimizer
                            result["type"] = texts["con_daily"]
                            result["nutrient"] = "MP"
                            result["unit"] = "g/day"
                            
                            if predicted_mp_g is not None and target is not None:
                                achieved = predicted_mp_g
                                result["actual"] = round(achieved, 0)
                                
                                tolerance_factor = tolerance_percent / 100.0
                                target_min = target * (1 - tolerance_factor)
                                target_max = target * (1 + tolerance_factor)
                                result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}% ({target_min:.0f} - {target_max:.0f})"
                                satisfied = target_min <= achieved <= target_max
                            else:
                                result["actual"] = texts["daily_intake_needed"]
                                result["condition"] = texts["cannot_calculate"]
                                satisfied = False
                                
                            result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                        elif attribute == "me":
                            # Metabolizable Energy constraint - use predicted_me_mcal from optimizer
                            result["type"] = texts["con_daily"]
                            result["nutrient"] = "ME"
                            result["unit"] = "Mcal/day"
                            
                            if predicted_me_mcal is not None and target is not None:
                                achieved = predicted_me_mcal
                                result["actual"] = round(achieved, 2)
                                
                                tolerance_factor = tolerance_percent / 100.0
                                target_min = target * (1 - tolerance_factor)
                                target_max = target * (1 + tolerance_factor)
                                result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}% ({target_min:.1f} - {target_max:.1f})"
                                satisfied = target_min <= achieved <= target_max
                            else:
                                result["actual"] = texts["daily_intake_needed"]
                                result["condition"] = texts["cannot_calculate"]
                                satisfied = False
                                
                            result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                        else:
                            # Nutrient daily total constraint
                            result["type"] = texts["con_daily"]
                            result["nutrient"] = attribute
                            result["unit"] = texts["con_daily"] # Daily Intake
                            
                            if daily_intake_kg and target is not None:
                                nutrient_percent = nutrient_analysis.get(attribute, 0.0)
                                achieved = (nutrient_percent / 100) * daily_intake_kg
                                result["actual"] = round(achieved, 2)
                                
                                tolerance_factor = tolerance_percent / 100.0
                                target_min = target * (1 - tolerance_factor)
                                target_max = target * (1 + tolerance_factor)
                                result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}% ({target_min:.2f} - {target_max:.2f})"
                                satisfied = target_min <= achieved <= target_max
                            else:
                                result["actual"] = texts["daily_intake_needed"]
                                result["condition"] = texts["cannot_calculate"]
                                satisfied = False
                                
                            result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    elif constraint_type == "ratio":
                        numerator = constraint["numerator"]
                        denominator = constraint["denominator"]
                        min_ratio = constraint.get("min")
                        max_ratio = constraint.get("max")
                        
                        result["type"] = texts["con_ratio"]
                        result["nutrient"] = f"{numerator}:{denominator}"
                        result["unit"] = texts["con_ratio"] # Ratio
                        
                        num_content = nutrient_analysis.get(numerator, 0.0)
                        denom_content = nutrient_analysis.get(denominator, 0.0)
                        
                        if denom_content > 0:
                            achieved = num_content / denom_content
                            result["actual"] = round(achieved, 2)
                            
                            conditions = []
                            if min_ratio is not None:
                                conditions.append(f"≥ {min_ratio}")
                            if max_ratio is not None:
                                conditions.append(f"≤ {max_ratio}")
                            result["condition"] = " & ".join(conditions)
                            
                            satisfied = True
                            if min_ratio is not None and achieved < min_ratio:
                                satisfied = False
                            if max_ratio is not None and achieved > max_ratio:
                                satisfied = False
                        else:
                            result["actual"] = texts["denom_zero"]
                            result["condition"] = texts["cannot_calculate"]
                            satisfied = False
                        
                        result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    constraint_results.append(result)
                
                return constraint_results
            
            # Get constraint validation results
            constraint_results = validate_constraints()

            # ==================== Run NASEM Evaluation First (dairy_cow only) ====================
            model_output = None
            predicted_milk = 0.0
            nasem_warning = None  # Track NASEM errors for agent feedback
            
            # Use stored animal_params from state for NASEM evaluation
            if animal_type == "dairy_cow" and not animal_params:
                return Command(
                    update={"messages": [ToolMessage(
                        "No animal parameters found in session. Please use set_animal_params first to set: body_weight, milk_prod, dim, parity.",
                        tool_call_id=tool_call_id
                    )]}
                )
            
            if animal_type == "dairy_cow" and animal_params:
                try:
                    from services.nasem_service import get_nasem_service
                    
                    nasem_service = get_nasem_service()
                    
                    # Build diet composition using centralized helper (single source of truth)
                    try:
                        from services.nasem_service import NASEMService
                        diet_composition, predicted_dmi_kg_for_nasem = NASEMService.build_diet_from_formulation(current_formulation)
                    except ValueError as e:
                        nasem_warning = f"Diet conversion failed: {e}"
                        logger.warning(nasem_warning)
                        model_output = None
                        predicted_milk = 0.0
                    
                    if diet_composition and feed_database:
                        # Build animal input for NASEM using stored animal_params
                        nasem_animal_input = nasem_service.build_animal_input(
                            body_weight_kg=animal_params.get("body_weight", 650),
                            days_in_milk=animal_params.get("dim", 100),
                            parity=animal_params.get("parity", 2),
                            target_milk_kg=animal_params.get("milk_prod", 35),
                            milk_fat_percent=animal_params.get("milk_fat_pct", 3.5),
                            milk_protein_percent=animal_params.get("milk_protein_pct", 3.2),
                            days_pregnant=animal_params.get("days_pregnant", 0),
                            breed=animal_params.get("breed", "Holstein"),
                            target_dmi_kg=predicted_dmi_kg_for_nasem,  # Use optimizer's DMI
                            bcs=animal_params.get("bcs", 3.0)
                        )
                        
                        # Build feedbase dict for NASEM
                        feedbase_dict = {"feeds": feed_database}
                        
                        # Run NASEM evaluation with full output
                        # Offload to thread pool (CPU-intensive numpy/pandas work)
                        nasem_results = await asyncio.to_thread(
                            nasem_service.evaluate_diet,
                            feedbase=feedbase_dict,
                            diet_composition=diet_composition,
                            animal_input=nasem_animal_input,
                            return_full_output=True
                        )
                        
                        if nasem_results.get("status") == "success":
                            model_output = nasem_results.get("model_output")
                            # Extract predicted milk for profitability
                            if model_output:
                                predicted_milk = model_output.get_value("Mlk_Prod_comp") or 0.0
                        
                except Exception as e:
                    nasem_warning = f"NASEM evaluation failed: {e}"
                    logger.warning(nasem_warning)
                    model_output = None
                    predicted_milk = 0.0

            # Create Excel with tabs
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

                # ==================== SINGLE TAB: Summary (All sections combined) ====================
                # Layout: A-E: ingredients, nutrients, constraints | F-G: profitability | H+: notes
                # Constraints and NASEM summary are side-by-side at the bottom
                
                # Build main column data (A-E)
                main_rows = []
                
                # Header
                main_rows.append([texts["sheet_summary"], "", "", "", ""])
                main_rows.append([f"{texts['date']}:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", ""])
                main_rows.append(["", "", "", "", ""])
                
                # Section 1: Feed Formula Table with Costs
                main_rows.append([texts["ingredients"], "", "", "", ""])
                
                # Calculate total cost and prepare cost data
                total_cost = 0.0
                feed_rows = []
                for feed_name, feed_data in formulation_feeds.items():
                    sanitized_name = sanitize_feed_name(feed_name)
                    kg_per_day = feed_data.get("kg_per_day", 0)
                    percentage_dm = feed_data.get("percentage_dm", 0)
                    
                    # Get price from feed database (field is 'cost_per_kg')
                    price_per_kg = 0.0
                    if feed_name in feed_database:
                        price_per_kg = feed_database[feed_name].get("cost_per_kg", 0) or 0
                    
                    # Use placeholder for cost formula - will be =B*D in Excel
                    # We still calculate for display/validation but use formula in Excel
                    cost_per_day = kg_per_day * price_per_kg if kg_per_day and price_per_kg else 0
                    total_cost += cost_per_day
                    
                    feed_rows.append([
                        sanitized_name,
                        round(kg_per_day, 2) if kg_per_day else 0,  # Use 0 instead of N/A for formula
                        round(percentage_dm, 1) if percentage_dm else "N/A",
                        round(price_per_kg, 2) if price_per_kg else 0,  # Use 0 instead of - for formula
                        "__COST_FORMULA__"  # Placeholder for =B*D formula
                    ])
                
                # Header row for feed table
                feed_header_row = len(main_rows) + 1  # 1-indexed row number in Excel
                main_rows.append([texts["feed_name"], texts["amount"], texts["dm_percent"], 
                                 texts["price_per_kg"], texts["cost_per_day"]])
                
                feed_data_start_row = len(main_rows) + 1  # First feed data row
                main_rows.extend(feed_rows)
                feed_data_end_row = len(main_rows)  # Last feed data row
                
                # Total row - use placeholders for SUM formulas (will be replaced with Excel formulas)
                total_row = len(main_rows) + 1  # Row number for TOTAL
                main_rows.append(["TOTAL", "__AMOUNT_SUM__", "100%", "", "__COST_SUM__"])
                main_rows.append(["", "", "", "", ""])
                
                # Section 2: Key Nutrients
                main_rows.append([texts["key_nutrients"], "", "", "", ""])
                
                # Key nutrients - map display names to NASEM field names
                key_nutrient_map = {
                    "CP": "Fd_CP",
                    "NDF": "Fd_NDF",
                    "ADF": "Fd_ADF",
                    "Fat": "Fd_CFat",
                    "Ca": "Fd_Ca",
                    "P": "Fd_P"
                }
                for display_name, field_name in key_nutrient_map.items():
                    value = nutrient_analysis.get(field_name, "-")
                    if isinstance(value, (int, float)):
                        value = round(value, 2)
                    main_rows.append([display_name, f"{value}%", "", "", ""])
                
                # F:C ratio
                forage_dm = 0.0
                concentrate_dm = 0.0
                for feed_name, feed_data in formulation_feeds.items():
                    kg = feed_data.get("kg_per_day", 0) or 0
                    if feed_name in feed_database:
                        category = feed_database[feed_name].get("type", "").lower()
                        if "forage" in category or "silage" in category or "hay" in category:
                            forage_dm += kg
                        else:
                            concentrate_dm += kg
                
                if forage_dm + concentrate_dm > 0:
                    forage_pct = round((forage_dm / (forage_dm + concentrate_dm)) * 100)
                    concentrate_pct = 100 - forage_pct
                    main_rows.append([texts["fc_ratio"], f"{forage_pct}:{concentrate_pct}", "", "", ""])
                
                main_rows.append(["", "", "", "", ""])
                
                # Track where constraints start for side-by-side with NASEM
                constraints_start_row = len(main_rows)
                
                # Section 3: Constraints (in main column area, will be side by side with NASEM)
                main_rows.append([texts["sheet_constraints"], "", "", "", ""])
                main_rows.append([texts["constraint_type"], texts["nutrient"], texts["condition"], texts["actual"], texts["satisfaction"]])
                
                for result in constraint_results:
                    main_rows.append([
                        result["type"],
                        result["nutrient"],
                        result["condition"],
                        result["actual"],
                        result["satisfaction"]
                    ])
                
                main_rows.append(["", "", "", "", ""])
                
                # Feed constraints
                main_rows.append([texts["feed_constraints"], "", "", "", ""])
                if feed_constraints:
                    main_rows.append([texts["feed_name"], texts["min_percent"], texts["max_percent"], "", ""])
                    for feed_name, constraints in feed_constraints.items():
                        sanitized_name = sanitize_feed_name(feed_name)
                        min_val = constraints.get('min', '')
                        max_val = constraints.get('max', '')
                        main_rows.append([sanitized_name, min_val, max_val, "", ""])
                else:
                    main_rows.append([texts["no_feed_constraints"], "", "", "", ""])
                
                # Build profitability column data (G-H, starting at row 1)
                # Check if milk price was explicitly set
                has_milk_price = "milk_price_per_kg" in animal_params
                milk_price_value = animal_params.get("milk_price_per_kg") if has_milk_price else None
                
                profit_rows = []
                profit_rows.append([texts["profitability"], ""])
                profit_rows.append([texts["input_section"], ""])
                profit_rows.append([texts["herd_size"], 100])  # Default 100
                profit_rows.append([texts["milk_price"], milk_price_value if milk_price_value is not None else ""])
                profit_rows.append(["", ""])
                
                # Cost metrics - always show (don't depend on milk price)
                profit_rows.append([texts["cost_per_kg_dm"], "__COST_PER_KG_DM__"])  # Formula placeholder
                profit_rows.append([texts["cost_per_cow_day"], "__COST_PER_COW__"])  # Formula placeholder
                profit_rows.append([texts["nasem_predicted_milk"], round(predicted_milk, 2) if predicted_milk else 0])
                # Revenue/profit rows - only populate formulas when milk price is explicitly set
                profit_rows.append([texts["revenue_per_cow_day"], 0 if has_milk_price else ""])
                profit_rows.append([texts["profit_per_cow_day"], 0 if has_milk_price else ""])
                profit_rows.append(["", ""])
                profit_rows.append([texts["herd_profit_day"], 0 if has_milk_price else ""])
                profit_rows.append([texts["herd_profit_month"], 0 if has_milk_price else ""])
                
                # Build notes as single text (will be put in vertically merged cell F)
                # Join all lines with newline for the merged cell
                if isinstance(description, str) and description.strip():
                    # Fix double-escaped newlines (e.g., from JSON serialization)
                    # Convert literal \n to actual newlines
                    notes_text = description.replace('\\n', '\n')
                else:
                    notes_text = texts["notes_none"]
                
                # Build NASEM summary for side column (starting at constraints_start_row in column G-H)
                nasem_summary_rows = []
                if model_output is not None:
                    nasem_summary_rows.append([texts["nasem_production"], ""])
                    
                    # Get values using ModelOutput.get_value()
                    mlk_prod = model_output.get_value("Mlk_Prod_comp")
                    mp_allow = model_output.get_value("Mlk_Prod_MPalow")
                    ne_allow = model_output.get_value("Mlk_Prod_NEalow")
                    milk_fat = model_output.get_value("MlkFat_Milk")
                    milk_protein = model_output.get_value("MlkNP_Milk")
                    
                    # Limiting factor
                    try:
                        mp_val = float(mp_allow) if mp_allow else 999
                        ne_val = float(ne_allow) if ne_allow else 999
                        if mp_val < ne_val:
                            limiting_factor = "MP (protein)"
                        elif ne_val < mp_val:
                            limiting_factor = "NE (energy)"
                        else:
                            limiting_factor = "Balanced"
                    except:
                        limiting_factor = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_predicted_milk"], round(mlk_prod, 2) if mlk_prod else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_allow_milk"], round(mp_allow, 2) if mp_allow else "N/A"])
                    nasem_summary_rows.append([texts["nasem_ne_allow_milk"], round(ne_allow, 2) if ne_allow else "N/A"])
                    nasem_summary_rows.append([texts["nasem_limiting_factor"], limiting_factor])
                    nasem_summary_rows.append([texts["nasem_milk_fat"], round(milk_fat, 4) if milk_fat else "N/A"])
                    nasem_summary_rows.append([texts["nasem_milk_protein"], round(milk_protein, 4) if milk_protein else "N/A"])
                    nasem_summary_rows.append(["", ""])
                    
                    # Energy Balance
                    nasem_summary_rows.append([texts["nasem_energy_balance"], ""])
                    me_intake = model_output.get_value("An_MEIn")
                    me_required = model_output.get_value("Trg_MEuse")
                    try:
                        me_balance = round(float(me_intake) - float(me_required), 2)
                        me_balance_str = f"+{me_balance}" if me_balance >= 0 else str(me_balance)
                    except:
                        me_balance_str = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_me_intake"], round(me_intake, 2) if me_intake else "N/A"])
                    nasem_summary_rows.append([texts["nasem_me_required"], round(me_required, 2) if me_required else "N/A"])
                    nasem_summary_rows.append([texts["nasem_me_balance"], me_balance_str])
                    nasem_summary_rows.append(["", ""])
                    
                    # Protein Balance
                    nasem_summary_rows.append([texts["nasem_protein_balance"], ""])
                    mp_intake = model_output.get_value("An_MPIn_g")
                    mp_required = model_output.get_value("An_MPuse_g_Trg")
                    rdp_intake = model_output.get_value("An_RDPIn_g")
                    try:
                        mp_balance = round(float(mp_intake) - float(mp_required), 1)
                        mp_balance_str = f"+{mp_balance}" if mp_balance >= 0 else str(mp_balance)
                    except:
                        mp_balance_str = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_mp_intake"], round(mp_intake, 1) if mp_intake else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_required"], round(mp_required, 1) if mp_required else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_balance"], mp_balance_str])
                    nasem_summary_rows.append([texts["nasem_rdp_intake"], round(rdp_intake, 1) if rdp_intake else "N/A"])
                    nasem_summary_rows.append(["", ""])
                    
                    # Extract AA data for summary + detailed sheet
                    AA_TARGETS = {"Lys": 7.2, "Met": 2.5}  # NASEM 2021 targets (% of MP)
                    aa_lys_pct = None
                    aa_met_pct = None
                    aa_limiting_str = "N/A"
                    try:
                        abs_aa_mpp = model_output.get_value("Abs_AA_MPp")
                        if abs_aa_mpp is not None:
                            aa_lys_pct = round(float(abs_aa_mpp.get("Lys", 0)), 2)
                            aa_met_pct = round(float(abs_aa_mpp.get("Met", 0)), 2)
                            limiting = []
                            for aa, tgt in AA_TARGETS.items():
                                if float(abs_aa_mpp.get(aa, 0)) < tgt:
                                    limiting.append(aa)
                            aa_limiting_str = ", ".join(limiting) if limiting else texts["nasem_aa_none_limiting"]
                    except Exception as e:
                        logger.warning(f"Failed to extract AA summary: {e}")
                    
                    # Other Key Metrics (includes compact AA summary)
                    nasem_summary_rows.append([texts["nasem_other_metrics"], ""])
                    dcad = model_output.get_value("An_DCADmeq")
                    nasem_summary_rows.append([texts["nasem_dcad"], round(dcad, 1) if dcad else "N/A"])
                    nasem_summary_rows.append([texts["nasem_lys_mp"], f"{aa_lys_pct}% (≥7.2%)" if aa_lys_pct is not None else "N/A"])
                    nasem_summary_rows.append([texts["nasem_met_mp"], f"{aa_met_pct}% (≥2.5%)" if aa_met_pct is not None else "N/A"])
                    nasem_summary_rows.append([texts["nasem_aa_limiting"], aa_limiting_str])
                
                # Determine max rows needed
                max_rows = max(len(main_rows), len(profit_rows), constraints_start_row + len(nasem_summary_rows))
                
                # Track notes cell merge range for later formatting (starts at row 2 since row 1 is header)
                notes_merge_start_row = 2
                notes_merge_end_row = max_rows
                
                # Create combined data with proper positioning
                # Layout: A-E (main) | F (notes with header) | G-H (profitability/NASEM)
                # Borders will be added between sections in formatting phase
                combined_data = []
                for i in range(max_rows):
                    row = []
                    
                    # Columns A-E (main content)
                    if i < len(main_rows):
                        row.extend(main_rows[i])
                    else:
                        row.extend(["", "", "", "", ""])
                    
                    # Column F (notes - header in row 1, text in row 2 for merge)
                    if i == 0:
                        row.append(texts["sheet_notes"])  # Header
                    elif i == 1:
                        row.append(notes_text)  # Notes text starts at row 2
                    else:
                        row.append("")  # Empty for merge
                    
                    # Columns G-H (profitability at top, NASEM summary at constraints_start_row)
                    if i < len(profit_rows):
                        # Profitability section at top
                        row.extend(profit_rows[i])
                    elif i >= constraints_start_row and nasem_summary_rows:
                        # NASEM summary parallel to constraints
                        nasem_idx = i - constraints_start_row
                        if nasem_idx < len(nasem_summary_rows):
                            row.extend(nasem_summary_rows[nasem_idx])
                        else:
                            row.extend(["", ""])
                    else:
                        row.extend(["", ""])
                    
                    combined_data.append(row)
                
                # Write combined data
                combined_df = pd.DataFrame(combined_data)
                combined_df.to_excel(writer, sheet_name=texts["sheet_summary"], index=False, header=False)
                
                # ==================== NASEM Category Tabs (dairy_cow only) ====================
                # Note: model_output was obtained earlier before Excel creation
                
                # Create NASEM category tabs if we have full output
                if model_output is not None:
                    # Helper function to flatten nested dict to list of [key, value] rows
                    # Skips 0 values, None, and empty values for cleaner output
                    def flatten_dict(d, prefix="", skip_zeros=True):
                        rows = []
                        for k, v in d.items():
                            # Build display key (show nested path)
                            display_key = f"{prefix}.{k}" if prefix else k
                            
                            if isinstance(v, dict):
                                rows.extend(flatten_dict(v, display_key, skip_zeros))
                            elif isinstance(v, pd.DataFrame):
                                # Skip DataFrames (too complex for simple key-value)
                                continue
                            elif isinstance(v, (list, pd.Series)):
                                # Format arrays - skip if empty
                                if hasattr(v, 'tolist'):
                                    v = v.tolist()
                                if v and len(v) > 0 and len(str(v)) < 100:
                                    rows.append([display_key, str(v)])
                            elif v is not None:
                                # Skip zeros if requested
                                if skip_zeros and isinstance(v, (int, float)) and v == 0:
                                    continue
                                # Skip empty strings
                                if isinstance(v, str) and not v.strip():
                                    continue
                                # Round floats for cleaner display
                                if isinstance(v, float):
                                    v = round(v, 3)
                                rows.append([display_key, v])
                        return rows
                    


                    # NASEM category to sheet name mapping (NASEM-Summary is now in main tab)
                    nasem_categories = {
                        "Intakes": "NASEM-Intakes",
                        "Requirements": "NASEM-Requirements", 
                        "Production": "NASEM-Production",
                        "Excretion": "NASEM-Excretion",
                        "Efficiencies": "NASEM-Efficiencies"
                    }

                    
                    # Export each category as a separate tab
                    for category_name, sheet_name in nasem_categories.items():
                        try:
                            category_data = getattr(model_output, category_name, None)
                            if category_data is None:
                                continue
                            
                            tab_data = []
                            tab_data.append([sheet_name])
                            tab_data.append([f"{texts['date']}:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                            tab_data.append([])
                            tab_data.append(["Variable", "Value"])
                            
                            # Flatten the category dict
                            if isinstance(category_data, dict):
                                rows = flatten_dict(category_data)
                                tab_data.extend(rows)
                            
                            # Write tab
                            if len(tab_data) > 4:  # Only if we have data beyond headers
                                tab_df = pd.DataFrame(tab_data)
                                tab_df.to_excel(writer, sheet_name=sheet_name[:31], index=False, header=False)
                                
                        except Exception as e:
                            logger.warning(f"Failed to export NASEM {category_name}: {e}")
                            continue
                    
                    # ==================== NASEM-AA Detailed Sheet ====================
                    try:
                        abs_aa_g = model_output.get_value("Abs_AA_g")       # pd.Series: AA → g/day
                        abs_aa_mpp = model_output.get_value("Abs_AA_MPp")   # pd.Series: AA → % of MP
                        trg_abs_aa = model_output.get_value("Trg_AbsAA_g")  # pd.Series: AA → target g/day
                        
                        if abs_aa_g is not None and abs_aa_mpp is not None:
                            aa_sheet_name = texts.get("nasem_aa_sheet", "NASEM-AA")
                            aa_tab_data = []
                            aa_tab_data.append([aa_sheet_name, "", "", "", "", ""])
                            aa_tab_data.append([f"{texts['date']}:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", ""])
                            aa_tab_data.append(["", "", "", "", "", ""])
                            
                            # Table header
                            aa_tab_data.append([
                                texts["nasem_aa_name"],
                                texts["nasem_aa_absorbed"],
                                texts["nasem_aa_target"],
                                texts["nasem_aa_pct_mp"],
                                texts["nasem_aa_target_pct"],
                                texts["nasem_aa_status"]
                            ])
                            
                            # All essential AAs
                            EAA_LIST = ["Arg", "His", "Ile", "Leu", "Lys", "Met", "Phe", "Thr", "Trp", "Val"]
                            for aa_name in EAA_LIST:
                                absorbed_g = round(float(abs_aa_g.get(aa_name, 0)), 1)
                                pct_mp = round(float(abs_aa_mpp.get(aa_name, 0)), 2)
                                target_g = round(float(trg_abs_aa.get(aa_name, 0)), 1) if trg_abs_aa is not None else ""
                                
                                # Target % only for Lys/Met (NASEM 2021 benchmarks)
                                target_pct = AA_TARGETS.get(aa_name, "")
                                target_pct_str = f"{target_pct}%" if target_pct else ""
                                
                                # Status: only evaluate for AAs with targets
                                if aa_name in AA_TARGETS:
                                    is_limiting = pct_mp < AA_TARGETS[aa_name]
                                    status = texts["nasem_aa_deficient"] if is_limiting else texts["nasem_aa_adequate"]
                                else:
                                    status = ""
                                
                                aa_tab_data.append([aa_name, absorbed_g, target_g, f"{pct_mp}%", target_pct_str, status])
                            
                            # Write the AA sheet
                            aa_df = pd.DataFrame(aa_tab_data)
                            aa_df.to_excel(writer, sheet_name=aa_sheet_name[:31], index=False, header=False)
                    except Exception as e:
                        logger.warning(f"Failed to export NASEM-AA sheet: {e}")
                

                # Apply formatting
                workbook = writer.book
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


                # Define common styles
                title_font = Font(bold=True, size=16, color="FFFFFF")
                title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                section_font = Font(bold=True, size=12, color="FFFFFF")
                section_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                label_font = Font(bold=True)
                thin_side = Side(style="thin", color="B4C6E7")
                table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                # ==================== FORMAT Summary Tab (unified) ====================
                if texts["sheet_summary"] in workbook.sheetnames:
                    ws = workbook[texts["sheet_summary"]]
                    
                    # Define yellow fill for editable input cells
                    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    # Define thick border for section dividers
                    thick_side = Side(style="medium", color="1F4E79")

                    # Auto-adjust column widths for new layout
                    # A-E: main content | F: notes | G-H: profitability/NASEM
                    ws.column_dimensions['A'].width = 30  # Main labels
                    ws.column_dimensions['B'].width = 15  # Main values
                    ws.column_dimensions['C'].width = 12  # Condition/DM%
                    ws.column_dimensions['D'].width = 12  # Actual/Price
                    ws.column_dimensions['E'].width = 15  # Satisfaction/Cost
                    ws.column_dimensions['F'].width = 50  # Notes column (wide)
                    ws.column_dimensions['G'].width = 28  # Profitability/NASEM labels
                    ws.column_dimensions['H'].width = 15  # Profitability/NASEM values

                    # Format notes header (F1) and merge notes text (F2:F{max_rows})
                    notes_header_cell = ws.cell(row=1, column=6)
                    notes_header_cell.font = section_font
                    notes_header_cell.fill = section_fill
                    
                    if notes_merge_end_row > notes_merge_start_row:
                        ws.merge_cells(f'F{notes_merge_start_row}:F{notes_merge_end_row}')
                        notes_cell = ws.cell(row=notes_merge_start_row, column=6)
                        notes_cell.font = Font(size=10)
                        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # Add outer borders around each section
                    max_row = ws.max_row
                    
                    # Helper to apply border preserving existing borders
                    def apply_border(cell, **sides):
                        existing = cell.border
                        cell.border = Border(
                            left=sides.get('left', existing.left),
                            right=sides.get('right', existing.right),
                            top=sides.get('top', existing.top),
                            bottom=sides.get('bottom', existing.bottom)
                        )
                    
                    # Apply borders to all cells
                    for row_num in range(1, max_row + 1):
                        for col_num in range(1, 9):  # A-H
                            cell = ws.cell(row=row_num, column=col_num)
                            
                            # Determine which borders to apply
                            left = thick_side if col_num == 1 else None  # Left edge of main
                            right = None
                            top = thick_side if row_num == 1 else None  # Top edge
                            bottom = thick_side if row_num == max_row else None  # Bottom edge
                            
                            # Right dividers between sections
                            if col_num == 5:  # E - divider after main
                                right = thick_side
                            elif col_num == 6:  # F - divider after notes
                                right = thick_side
                            elif col_num == 8:  # H - right edge of profitability
                                right = thick_side
                            
                            apply_border(cell, left=left, right=right, top=top, bottom=bottom)

                    # Track special cells for formula insertion in column H (profitability values)
                    herd_size_cell = None
                    milk_price_cell = None
                    predicted_milk_cell = None
                    cost_per_cow_cell = None
                    profit_row = None
                    
                    # Apply formatting based on content - scan all columns
                    for row_num in range(1, ws.max_row + 1):
                        # Column A content (main area)
                        a_cell = ws.cell(row=row_num, column=1)
                        a_value = str(a_cell.value or "")
                        b_cell = ws.cell(row=row_num, column=2)
                        
                        # Column G content (profitability/NASEM labels)
                        g_cell = ws.cell(row=row_num, column=7)
                        g_value = str(g_cell.value or "")
                        h_cell = ws.cell(row=row_num, column=8)

                        # ===== Main area formatting (column A) =====
                        # Title row
                        if a_value == texts["sheet_summary"]:
                            a_cell.font = title_font
                            a_cell.fill = title_fill
                            ws.merge_cells(f'A{row_num}:E{row_num}')

                        # Section headers in main area
                        elif a_value in [texts["ingredients"], texts["key_nutrients"], 
                                        texts["sheet_constraints"], texts["feed_constraints"]]:
                            a_cell.font = section_font
                            a_cell.fill = section_fill
                            ws.merge_cells(f'A{row_num}:E{row_num}')

                        # Feed table header
                        elif a_value == texts["feed_name"]:
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")
                        
                        # Constraint table header
                        elif a_value == texts["constraint_type"]:
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")

                        # TOTAL row - apply formatting and insert SUM formulas for columns B and E
                        elif a_value == "TOTAL":
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                cell.font = Font(bold=True)
                            # Replace placeholder with SUM formula for amount (column B)
                            b_cell_total = ws.cell(row=row_num, column=2)
                            if str(b_cell_total.value) == "__AMOUNT_SUM__":
                                b_cell_total.value = f"=SUM(B{feed_data_start_row}:B{feed_data_end_row})"
                            # Replace placeholder with SUM formula for feed costs (column E)
                            e_cell_total = ws.cell(row=row_num, column=5)
                            if str(e_cell_total.value) == "__COST_SUM__":
                                e_cell_total.value = f"=SUM(E{feed_data_start_row}:E{feed_data_end_row})"

                        # Label rows in main area (ending with :)
                        elif a_value.endswith(":"):
                            a_cell.font = label_font
                        
                        # Replace __COST_FORMULA__ placeholder with =B*D formula for feed cost rows
                        e_cell = ws.cell(row=row_num, column=5)
                        e_value = str(e_cell.value or "")
                        if e_value == "__COST_FORMULA__":
                            # Cost = Amount (B) × Price (D)
                            e_cell.value = f"=B{row_num}*D{row_num}"
                        
                        # Satisfaction pass/fail coloring in column E
                        e_cell = ws.cell(row=row_num, column=5)
                        e_value = str(e_cell.value or "")
                        if texts["satisfied"] in e_value:
                            e_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            e_cell.font = Font(color="006100")
                        elif texts["unsatisfied"] in e_value:
                            e_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            e_cell.font = Font(color="9C0006")

                        # ===== G-H area formatting (profitability + NASEM) =====
                        # Profitability/NASEM section headers (now in column G)
                        if g_value in [texts["profitability"], texts["nasem_production"], 
                                      texts["nasem_energy_balance"], texts["nasem_protein_balance"],
                                      texts["nasem_other_metrics"]]:
                            g_cell.font = section_font
                            g_cell.fill = section_fill
                            ws.merge_cells(f'G{row_num}:H{row_num}')

                        # Input cells - highlight yellow and track for formulas (values now in H)
                        elif g_value == texts["herd_size"]:
                            h_cell.fill = input_fill
                            herd_size_cell = f"H{row_num}"
                        elif g_value == texts["milk_price"]:
                            h_cell.fill = input_fill
                            milk_price_cell = f"H{row_num}"
                        
                        # Track cells for formulas
                        elif g_value == texts["nasem_predicted_milk"]:
                            predicted_milk_cell = f"H{row_num}"
                        
                        # Cost per kg DM - formula referencing E total / B total (DMI)
                        elif g_value == texts["cost_per_kg_dm"]:
                            h_cell_value = str(h_cell.value or "")
                            if h_cell_value == "__COST_PER_KG_DM__":
                                # Cost per kg DM = Total Cost (E{total_row}) / DMI (B{total_row})
                                h_cell.value = f"=E{total_row}/B{total_row}"
                        
                        # Cost per cow per day - formula referencing E total
                        elif g_value == texts["cost_per_cow_day"]:
                            h_cell_value = str(h_cell.value or "")
                            if h_cell_value == "__COST_PER_COW__":
                                # Cost per cow = Total Cost from column E
                                h_cell.value = f"=E{total_row}"
                            cost_per_cow_cell = f"H{row_num}"
                        
                        # Revenue formula - only apply when milk price is explicitly set
                        elif g_value == texts["revenue_per_cow_day"]:
                            if has_milk_price:
                                # Revenue = Predicted Milk × Milk Price
                                if predicted_milk_cell and milk_price_cell:
                                    h_cell.value = f"={predicted_milk_cell}*{milk_price_cell}"
                                else:
                                    h_cell.value = f"=H{row_num-1}*H{row_num-5}"
                        
                        # Profit per cow formula - only apply when milk price is explicitly set
                        elif g_value == texts["profit_per_cow_day"]:
                            if has_milk_price:
                                if predicted_milk_cell and milk_price_cell and cost_per_cow_cell:
                                    h_cell.value = f"={predicted_milk_cell}*{milk_price_cell}-{cost_per_cow_cell}"
                                else:
                                    h_cell.value = f"=H{row_num-1}-H{row_num-3}"
                            profit_row = row_num
                        
                        # Herd profit/day formula - only apply when milk price is explicitly set
                        elif g_value == texts["herd_profit_day"]:
                            if has_milk_price and profit_row and herd_size_cell:
                                h_cell.value = f"=H{profit_row}*{herd_size_cell}"
                        
                        # Herd profit/month formula - only apply when milk price is explicitly set
                        elif g_value == texts["herd_profit_month"]:
                            if has_milk_price:
                                herd_day_cell = f"H{row_num - 1}" if row_num > 1 else None
                                if herd_day_cell:
                                    h_cell.value = f"={herd_day_cell}*30"

                        # Input section label
                        elif g_value == texts["input_section"]:
                            g_cell.font = Font(italic=True, color="666666")

                        # Label rows in G column (ending with :)
                        elif g_value.endswith(":"):
                            g_cell.font = label_font

                # ==================== FORMAT NASEM Category Tabs ====================
                # Note: NASEM-Summary is now in main tab, so not included here
                nasem_sheet_names = ["NASEM-Intakes", "NASEM-Requirements", "NASEM-Production", 
                                     "NASEM-Excretion", "NASEM-Efficiencies"]
                
                for sheet_name in nasem_sheet_names:
                    if sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        
                        # Set column widths
                        ws.column_dimensions['A'].width = 50
                        ws.column_dimensions['B'].width = 25
                        
                        for row_num in range(1, ws.max_row + 1):
                            first_cell = ws.cell(row=row_num, column=1)
                            first_value = str(first_cell.value or "")
                            
                            # Title row (first row with sheet name)
                            if first_value == sheet_name:
                                first_cell.font = title_font
                                first_cell.fill = title_fill
                                ws.merge_cells(f'A{row_num}:B{row_num}')
                            
                            # Header row
                            elif first_value == "Variable":
                                first_cell.font = header_font
                                first_cell.fill = header_fill
                                second_cell = ws.cell(row=row_num, column=2)
                                if second_cell.value:
                                    second_cell.font = header_font
                                    second_cell.fill = header_fill
                            
                            # Label rows (ending with :)
                            elif first_value.endswith(":"):
                                first_cell.font = label_font
                
                # ==================== FORMAT NASEM-AA Sheet ====================
                aa_sheet_name = texts.get("nasem_aa_sheet", "NASEM-AA")
                if aa_sheet_name in workbook.sheetnames:
                    ws = workbook[aa_sheet_name]
                    
                    # Column widths for 6-column table
                    ws.column_dimensions['A'].width = 18  # AA Name
                    ws.column_dimensions['B'].width = 16  # Absorbed
                    ws.column_dimensions['C'].width = 16  # Target
                    ws.column_dimensions['D'].width = 14  # % MP
                    ws.column_dimensions['E'].width = 14  # Target %
                    ws.column_dimensions['F'].width = 16  # Status
                    
                    for row_num in range(1, ws.max_row + 1):
                        first_cell = ws.cell(row=row_num, column=1)
                        first_value = str(first_cell.value or "")
                        
                        # Title row
                        if first_value == aa_sheet_name:
                            first_cell.font = title_font
                            first_cell.fill = title_fill
                            ws.merge_cells(f'A{row_num}:F{row_num}')
                        
                        # Header row (AA Name column)
                        elif first_value == texts["nasem_aa_name"]:
                            for col in range(1, 7):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")
                        
                        # Data rows - format status column with pass/fail colors
                        elif first_value in ["Arg", "His", "Ile", "Leu", "Lys", "Met", "Phe", "Thr", "Trp", "Val"]:
                            status_cell = ws.cell(row=row_num, column=6)
                            status_val = str(status_cell.value or "")
                            if texts["nasem_aa_adequate"] in status_val:
                                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                status_cell.font = Font(color="006100")
                            elif texts["nasem_aa_deficient"] in status_val:
                                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                status_cell.font = Font(color="9C0006")
                            # Center-align data cells
                            for col in range(2, 7):
                                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
                        
                        # Date row
                        elif first_value.endswith(":"):
                            first_cell.font = label_font
            


            # Format file info for backend parsing
            # Convert literal \n from LLM to actual newlines
            normalized_description = description.replace('\\n', '\n') if description else None

            file_info = {
                "filepath": str(filepath),
                "filename": filename,
                "type": "excel",
                "description": normalized_description
            }

            # Build success message with warnings if any
            success_msg = texts["success"].format(filename=filename)
            if nasem_warning:
                success_msg += f" ⚠️ {nasem_warning}"
            success_msg += f" [FILE_EXPORT]{json.dumps(file_info, ensure_ascii=False)}[/FILE_EXPORT]"

            return Command(
                update={"messages": [ToolMessage(success_msg, tool_call_id=tool_call_id)]}
            )
            
        except Exception as e:
            logger.error(f"Export formulation error: {e}")
            return Command(
                update={"messages": [ToolMessage(texts["fail"].format(error=str(e)), tool_call_id=tool_call_id)]}
            )
    return export_formulation
