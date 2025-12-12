import asyncio
import json
import logging
import re
from typing import Dict, List, Any, Optional, Annotated
from datetime import datetime
from pathlib import Path
import pandas as pd
from langchain_core.tools import tool, InjectedToolCallId, InjectedToolArg
from langchain_core.messages import ToolMessage
from langchain_core.runnables import RunnableConfig
from langgraph.prebuilt import InjectedState
from langgraph.types import Command
from langgraph.config import get_store

from formulation.optimizer import create_optimizer
from utils.language import normalize_locale

logger = logging.getLogger(__name__)

# Translation dictionary for export
EXPORT_TRANSLATIONS = {
    "zh-CN": {
        "sheet_results": "配方结果",
        "sheet_notes": "配方说明",
        "sheet_constraints": "约束条件",
        "date": "导出日期",
        "status": "优化状态",
        "total_cost": "总成本",
        "cost_unit": "元/公斤干物质",
        "dmi": "日干物质采食量",
        "ingredients": "饲料配方明细",
        "feed_name": "饲料名称",
        "amount": "日饲喂量 (kg/day)",
        "dm_percent": "干物质比例 (%)",
        "nutrients_header": "原料含量 (% DM)",
        "nutrition_profile": "整体营养成分分析",
        "nutrient": "营养成分",
        "content": "含量 (% DM)",
        "notes_none": "暂无配方说明",
        "constraint_validation": "营养约束验证",
        "constraint_type": "约束类型",
        "condition": "约束条件",
        "actual": "实际值",
        "unit": "单位",
        "satisfaction": "满足情况",
        "feed_constraints": "饲料用量约束",
        "min_percent": "最小比例 (%)",
        "max_percent": "最大比例 (%)",
        "no_feed_constraints": "无饲料用量约束",
        "chart_title": "营养成分组成图",
        "satisfied": "✓ 满足",
        "unsatisfied": "✗ 不满足",
        "con_concentration": "浓度约束",
        "con_dmi": "干物质采食量约束",
        "con_daily": "日摄入约束",
        "con_ratio": "比例约束",
        "target": "目标",
        "range": "范围内",
        "daily_intake_needed": "需要日采食量",
        "cannot_calculate": "无法计算",
        "denom_zero": "分母为零",
        "error_no_formulation": "未找到成功的配方。请先运行配方优化。",
        "error_no_db": "未找到饲料数据库。导出可能不包含完整的饲料信息。",
        "export_success": "✅ 成功导出 {filename}。",
        "export_fail": "导出配方时出错: {error}",
        # NASEM sheet translations
        "sheet_nasem": "NASEM分析",
        "nasem_production": "产奶预测",
        "nasem_predicted_milk": "预测产奶量 (kg/day)",
        "nasem_mp_allow_milk": "MP允许产奶量 (kg)",
        "nasem_ne_allow_milk": "NE允许产奶量 (kg)",
        "nasem_limiting_factor": "限制因素",
        "nasem_milk_fat": "乳脂 (g/g)",
        "nasem_milk_protein": "乳蛋白 (g/g)",
        "nasem_energy_balance": "能量平衡",
        "nasem_me_intake": "ME摄入 (Mcal/d)",
        "nasem_me_required": "ME需求 (Mcal/d)",
        "nasem_me_balance": "ME平衡",
        "nasem_protein_balance": "蛋白质平衡",
        "nasem_mp_intake": "MP摄入 (g/d)",
        "nasem_mp_required": "MP需求 (g/d)",
        "nasem_mp_balance": "MP平衡",
        "nasem_rdp_intake": "RDP摄入 (g/d)",
        "nasem_other_metrics": "其他指标",
        "nasem_dmi": "干物质采食量 (kg)",
        "nasem_dcad": "DCAD (meq)",
        "nasem_lys_mp": "赖氨酸占MP比例 (%)",
        "nasem_met_mp": "蛋氨酸占MP比例 (%)",
        "nasem_no_animal_input": "未提供动物参数,跳过NASEM分析",
        # Profitability section
        "sheet_summary": "配方摘要",
        "key_nutrients": "关键营养指标",
        "profitability": "经济效益分析",
        "input_section": "输入参数 (黄色单元格可编辑)",
        "herd_size": "牛群规模 (头)",
        "milk_price": "奶价 (元/kg)",
        "cost_per_kg_dm": "成本 (元/kg DM)",
        "cost_per_cow_day": "每头日成本 (元)",
        "revenue_per_cow_day": "每头日收入 (元)",
        "profit_per_cow_day": "每头日利润 (元)",
        "herd_profit_day": "牛群日利润 (元)",
        "herd_profit_month": "牛群月利润 (元)",
        "fc_ratio": "粗精比",
        "forage": "粗料",
        "concentrate": "精料",
        "price_per_kg": "价格 (元/kg)",
        "cost_per_day": "日成本 (元)"
    },
    "en-US": {
        "sheet_results": "Formulation Results",
        "sheet_notes": "Formulation Notes",
        "sheet_constraints": "Constraints",
        "date": "Date",
        "status": "Status",
        "total_cost": "Total Cost",
        "cost_unit": "CNY/kg DM",
        "dmi": "DMI",
        "ingredients": "Ingredients",
        "feed_name": "Feed Name",
        "amount": "Amount (kg/day)",
        "dm_percent": "DM %",
        "nutrients_header": "Nutrients (% DM)",
        "nutrition_profile": "Nutrition Profile",
        "nutrient": "Nutrient",
        "content": "Content (% DM)",
        "notes_none": "No notes available",
        "constraint_validation": "Nutrient Constraints Validation",
        "constraint_type": "Type",
        "condition": "Condition",
        "actual": "Actual",
        "unit": "Unit",
        "satisfaction": "Status",
        "feed_constraints": "Feed Constraints",
        "min_percent": "Min %",
        "max_percent": "Max %",
        "no_feed_constraints": "No feed constraints",
        "chart_title": "Nutrition Composition",
        "satisfied": "✓ Pass",
        "unsatisfied": "✗ Fail",
        "con_concentration": "Concentration",
        "con_dmi": "DMI Constraint",
        "con_daily": "Daily Intake",
        "con_ratio": "Ratio",
        "target": "Target",
        "range": "In range",
        "daily_intake_needed": "Daily intake needed",
        "cannot_calculate": "Cannot calculate",
        "denom_zero": "Denominator zero",
        "error_no_formulation": "No successful formulation found. Please run optimization first.",
        "error_no_db": "Feed database not found. Export may not include complete feed information.",
        "export_success": "✅ Successfully exported {filename}.",
        "export_fail": "Error exporting formulation: {error}",
        # NASEM sheet translations
        "sheet_nasem": "NASEM Analysis",
        "nasem_production": "Production Predictions",
        "nasem_predicted_milk": "Predicted Milk (kg/day)",
        "nasem_mp_allow_milk": "MP Allowable Milk (kg)",
        "nasem_ne_allow_milk": "NE Allowable Milk (kg)",
        "nasem_limiting_factor": "Limiting Factor",
        "nasem_milk_fat": "Milk Fat (g/g)",
        "nasem_milk_protein": "Milk Protein (g/g)",
        "nasem_energy_balance": "Energy Balance",
        "nasem_me_intake": "ME Intake (Mcal/d)",
        "nasem_me_required": "ME Required (Mcal/d)",
        "nasem_me_balance": "ME Balance",
        "nasem_protein_balance": "Protein Balance",
        "nasem_mp_intake": "MP Intake (g/d)",
        "nasem_mp_required": "MP Required (g/d)",
        "nasem_mp_balance": "MP Balance",
        "nasem_rdp_intake": "RDP Intake (g/d)",
        "nasem_other_metrics": "Other Metrics",
        "nasem_dmi": "DMI (kg)",
        "nasem_dcad": "DCAD (meq)",
        "nasem_lys_mp": "Lys % of MP",
        "nasem_met_mp": "Met % of MP",
        "nasem_no_animal_input": "No animal input provided, NASEM analysis skipped",
        # Profitability section
        "sheet_summary": "Summary",
        "key_nutrients": "Key Nutrients",
        "profitability": "Profitability Analysis",
        "input_section": "Inputs (edit yellow cells)",
        "herd_size": "Herd Size",
        "milk_price": "Milk Price (¥/kg)",
        "cost_per_kg_dm": "Cost (¥/kg DM)",
        "cost_per_cow_day": "Cost/Cow/Day (¥)",
        "revenue_per_cow_day": "Revenue/Cow/Day (¥)",
        "profit_per_cow_day": "Profit/Cow/Day (¥)",
        "herd_profit_day": "Herd Profit/Day (¥)",
        "herd_profit_month": "Herd Profit/Month (¥)",
        "fc_ratio": "F:C Ratio",
        "forage": "Forage",
        "concentrate": "Concentrate",
        "price_per_kg": "Price (¥/kg)",
        "cost_per_day": "Cost/Day (¥)"
    }
}


def sanitize_feed_name(name: str) -> str:
    """
    Sanitize feed names for Excel export by removing control characters and null bytes.
    
    Args:
        name: Original feed name
        
    Returns:
        Sanitized feed name safe for Excel
    """
    if not isinstance(name, str):
        return str(name)
    
    # Remove control characters (0x00-0x1F and 0x7F-0x9F)
    sanitized = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', name)
    
    # Remove any remaining null bytes
    sanitized = sanitized.replace('\x00', '')
    
    # Strip whitespace
    sanitized = sanitized.strip()
    
    # If name becomes empty after sanitization, provide a default
    if not sanitized:
        sanitized = "饲料"
    
    return sanitized


def create_formulation_tools(animal_type: str = "dairy_cow"):
    """Create formulation tools that operate on LangGraph state.

    Args:
        animal_type: Animal type for feedbase filtering (dairy_cow, beef_cow, cat, dog)
    """
    
    def _is_free_tier(config: Optional[RunnableConfig]) -> bool:
        """Determine whether current LangGraph run belongs to a free tier account."""
        try:
            tier = (config or {}).get("configurable", {}).get("account_tier")
            return (tier or "free") == "free"
        except AttributeError:
            return True

    def _free_tier_feedbase_message(action: str, tool_call_id: Optional[str]) -> Command:
        """Standardized response when free accounts attempt custom feedbase actions."""
        return Command(
            update={
                "messages": [
                    ToolMessage(
                        f"Free tier accounts can only {action} system feedbases named default_*. Please upgrade to unlock custom feedbases.",
                        tool_call_id=tool_call_id,
                    )
                ]
            }
        )
    
    @tool
    async def add_feed(
        feed_base_name: str,
        name: str,
        dm_percent: Optional[float] = None,
        dry_matter_percent: Optional[float] = None,
        nutrients: Dict[str, float] = None,
        cost_per_kg: float = 0.0,
        state: Annotated[dict, InjectedState] = None,
        tool_call_id: Annotated[str, InjectedToolCallId] = None,
        config: RunnableConfig = None
    ) -> Command:
        """
        Add or update feed ingredient in a specific feedbase.

        Args:
            feed_base_name: Name of the feedbase to add the feed to(will create if not exists)
            name: Feed name (will replace if exists)
            dm_percent: Dry matter percentage (0-100)
            nutrients: Nutrient composition on dry matter basis (e.g., {"CP": 18.5, "NEL": 1.65})
            cost_per_kg: Cost per kg as-fed

        Returns:
            Confirmation message
        """
        try:
            # Validate inputs
            if not isinstance(feed_base_name, str) or not feed_base_name.strip():
                return Command(
                    update={
                        "messages": [
                            ToolMessage("Error: Feedbase name must be a non-empty string", tool_call_id=tool_call_id)
                        ]
                    }
                )

            if _is_free_tier(config):
                return _free_tier_feedbase_message("create or edit", tool_call_id)

            # Block creating feedbases starting with "default"
            if feed_base_name.startswith("default"):
                return Command(
                    update={
                        "messages": [
                            ToolMessage("Error: Feedbase names cannot start with 'default'. This prefix is reserved for system feedbases.", tool_call_id=tool_call_id)
                        ]
                    }
                )
                
            if not isinstance(name, str) or not name.strip():
                return Command(
                    update={
                        "messages": [
                            ToolMessage("Error: Feed name must be a non-empty string", tool_call_id=tool_call_id)
                        ]
                    }
                )
            
            # Accept legacy param name
            if dm_percent is None and dry_matter_percent is not None:
                dm_percent = dry_matter_percent
            if dm_percent is None:
                return Command(
                    update={"messages": [ToolMessage("Error: dm_percent (dry matter %) is required", tool_call_id=tool_call_id)]}
                )
            if not 0 < dm_percent <= 100:
                return Command(
                    update={"messages": [ToolMessage("Error: Dry matter percentage must be between 0 and 100", tool_call_id=tool_call_id)]}
                )
            
            if not isinstance(nutrients, dict) or not nutrients:
                return Command(
                    update={"messages": [ToolMessage("Error: Nutrients must be a non-empty dictionary", tool_call_id=tool_call_id)]}
                )
            
            if cost_per_kg < 0:
                return Command(
                    update={"messages": [ToolMessage("Error: Cost per kg must be non-negative", tool_call_id=tool_call_id)]}
                )
            
            # Validate nutrient values
            for nutrient, value in nutrients.items():
                if not isinstance(value, (int, float)) or value < 0:
                    return Command(
                        update={"messages": [ToolMessage(f"Error: Nutrient '{nutrient}' must have a non-negative numeric value", tool_call_id=tool_call_id)]}
                    )
            
            # Get user_id from config and access store
            user_id = config["configurable"].get("user_id")
            if not user_id:
                return Command(
                    update={"messages": [ToolMessage("Error: User ID not found in configuration", tool_call_id=tool_call_id)]}
                )
            
            store = get_store()
            namespace = ("feedbases", user_id, feed_base_name)
            
            # Get existing feedbase or create new one
            existing_feedbase = await store.aget(namespace, "data")
            if existing_feedbase:
                feedbase_data = existing_feedbase.value
            else:
                # Create new feedbase with animal_type metadata
                feedbase_data = {
                    "animal_type": animal_type,
                    "feeds": {}
                }
            
            # Format feed data
            feed_data = {
                "dm_percent": dm_percent,
                "nutrients": nutrients,
                "cost_per_kg": cost_per_kg
            }
            
            # Add feed to feedbase
            sanitized_name = sanitize_feed_name(name)
            feedbase_data["feeds"].update({sanitized_name: feed_data})
            
            # Store updated feedbase
            await store.aput(namespace, "data", feedbase_data)
            
            # Return success message
            return Command(
                update={"messages": [ToolMessage(f"Successfully added feed '{sanitized_name}' to feedbase '{feed_base_name}'", tool_call_id=tool_call_id)]}
            )
            
        except Exception as e:
            logger.error(f"Add feed error: {e}")
            return Command(
                update={"messages": [ToolMessage(f"Error adding feed: {str(e)}", tool_call_id=tool_call_id)]}
            )
    
    @tool
    async def list_feed_bases(
        state: Annotated[dict, InjectedState],
        tool_call_id: Annotated[str, InjectedToolCallId],
        config: RunnableConfig
    ) -> Command:
        """
        List all available feedbases for the user filtered by animal type.
        Includes both user-created feedbases and system default feedbases.

        Returns:
            List of available feedbase names matching the current animal type
        """
        try:
            # Get user_id from config and access store
            user_id = config["configurable"].get("user_id")
            if not user_id:
                return Command(
                    update={"messages": [ToolMessage("Error: User ID not found in configuration", tool_call_id=tool_call_id)]}
                )

            free_tier = _is_free_tier(config)
            store = get_store()

            if free_tier:
                system_feedbase_name = f"default_{animal_type}"
                system_namespace = ("system_feedbases", system_feedbase_name)
                system_feedbase = await store.aget(system_namespace, "data")

                if not system_feedbase:
                    return Command(
                        update={"messages": [ToolMessage(f"No system feedbase found for animal type '{animal_type}'. Please contact support.", tool_call_id=tool_call_id)]}
                    )

                feed_count = len(system_feedbase.value.get("feeds", {}))
                message_lines = [
                    f"Available Feedbases for {animal_type} (1):",
                    "**System Feedbases:**",
                    f"- **{system_feedbase_name}** ({feed_count} feeds) [READ-ONLY]",
                    "",
                    "Upgrade your plan to create and manage custom feedbases."
                ]
                return Command(
                    update={"messages": [ToolMessage("\n".join(message_lines), tool_call_id=tool_call_id)]}
                )

            # Search for user feedbases
            user_namespace = ("feedbases", user_id)
            user_feedbase_entries = await store.asearch(user_namespace)

            # Search for system default feedbase for this animal type
            system_feedbase_name = f"default_{animal_type}"
            system_namespace = ("system_feedbases", system_feedbase_name)
            system_feedbase = await store.aget(system_namespace, "data")

            # Filter user feedbases by animal_type
            filtered_feedbases = []
            for entry in user_feedbase_entries:
                feedbase_data = entry.value
                feedbase_animal_type = feedbase_data.get("animal_type", "dairy_cow")  # Default to dairy_cow for legacy data

                if feedbase_animal_type == animal_type:
                    filtered_feedbases.append(entry)

            # Build response
            feedbase_info = []
            total_count = len(filtered_feedbases)

            # Add system default feedbase if it exists and has feeds
            if system_feedbase and len(system_feedbase.value.get("feeds", {})) > 0:
                total_count += 1
                feedbase_info.append(f"Available Feedbases for {animal_type} ({total_count}):\n")
                feedbase_info.append("**System Feedbases:**")
                feed_count = len(system_feedbase.value.get("feeds", {}))
                feedbase_info.append(f"- **{system_feedbase_name}** ({feed_count} feeds) [READ-ONLY]")
                feedbase_info.append("")
            else:
                feedbase_info.append(f"Available Feedbases for {animal_type} ({total_count}):\n")

            # Add user feedbases
            if filtered_feedbases:
                if system_feedbase and len(system_feedbase.value.get("feeds", {})) > 0:
                    feedbase_info.append("**User Feedbases:**")
                for entry in filtered_feedbases:
                    # Namespace structure: ("feedbases", user_id, feedbase_name)
                    feedbase_name = entry.namespace[2]  # Get feedbase name from namespace
                    feedbase_data = entry.value
                    feed_count = len(feedbase_data.get("feeds", {}))
                    feedbase_info.append(f"- **{feedbase_name}** ({feed_count} feeds)")

            if total_count == 0:
                return Command(
                    update={"messages": [ToolMessage(f"No feedbases found for animal type '{animal_type}'.", tool_call_id=tool_call_id)]}
                )

            return Command(
                update={"messages": [ToolMessage("\n".join(feedbase_info), tool_call_id=tool_call_id)]}
            )

        except Exception as e:
            logger.error(f"List feedbases error: {e}")
            return Command(
                update={"messages": [ToolMessage(f"Error listing feedbases: {str(e)}", tool_call_id=tool_call_id)]}
            )
    
    @tool
    async def formulate_ration(
        feed_base_name: str,
        nutritional_constraints: List[Dict[str, Any]],
        selected_feeds: List[str],
        state: Annotated[dict, InjectedState],
        tool_call_id: Annotated[str, InjectedToolCallId],
        config: RunnableConfig,
        feed_constraints: Optional[Dict[str, Dict]] = None,
        optimization_goal: str = "minimize_cost"
    ) -> Command:
        """
        Formulate optimal ration using flexible constraint system with tolerance support.
        Automatically calculates daily feed amounts when DMI is specified in constraints.
        
        Args:
            feed_base_name: Name of the feedbase to use for formulation
            nutritional_constraints: List of constraint dictionaries:
                - {"type": "concentration", "nutrient": "CP", "min": 16.0, "max": 18.0}
                - {"type": "daily_total", "attribute": "dmi", "target": 18.0, "tolerance_percent": 10.0}
                - {"type": "daily_total", "attribute": "NEL", "target": 32.0, "tolerance_percent": 5.0}
                - {"type": "daily_total", "attribute": "CP", "target": 2.8, "tolerance_percent": 8.0}
                - {"type": "ratio", "numerator": "Ca", "denominator": "P", "min": 1.2, "max": 2.0}
            selected_feeds: List of feed names to include in optimization
            feed_constraints: Optional inclusion limits {"feed_name": {"min": 0, "max": 50}}
            optimization_goal: "minimize_cost" or other objectives
            
        Daily Total Constraints:
            - Use "daily_total" type for any daily target (DMI, nutrients, etc.)
            - "attribute": can be "dmi" or any nutrient name (e.g., "NEL", "CP", "Ca")
            - "target": the target daily amount
            - "tolerance_percent": optional tolerance (default 10%) - creates flexible range around target
            - DMI constraint should be specified first, as nutrient daily totals depend on it
            
        Returns:
            State update with formulation results including daily amounts when DMI is provided
        """
        try:
            # Validate inputs
            if not isinstance(feed_base_name, str) or not feed_base_name.strip():
                return Command(
                    update={"messages": [ToolMessage("Error: Feedbase name must be a non-empty string", tool_call_id=tool_call_id)]}
                )
            
            if _is_free_tier(config) and not feed_base_name.startswith("default_"):
                return _free_tier_feedbase_message("use custom", tool_call_id)
                
            if not isinstance(nutritional_constraints, list):
                return Command(
                    update={"messages": [ToolMessage("Error: nutritional_constraints must be a list", tool_call_id=tool_call_id)]}
                )
            
            if not isinstance(selected_feeds, list) or not selected_feeds:
                return Command(
                    update={"messages": [ToolMessage("Error: selected_feeds must be a non-empty list", tool_call_id=tool_call_id)]}
                )
            
            # Validate constraint format
            for i, constraint in enumerate(nutritional_constraints):
                if not isinstance(constraint, dict):
                    return Command(
                        update={"messages": [ToolMessage(f"Error: Constraint {i} must be a dictionary", tool_call_id=tool_call_id)]}
                    )
                
                constraint_type = constraint.get("type", "")
                if constraint_type not in ["concentration", "daily_total", "ratio"]:
                    return Command(
                        update={"messages": [ToolMessage(f"Error: Invalid constraint type '{constraint_type}' in constraint {i}. Must be 'concentration', 'daily_total', or 'ratio'", tool_call_id=tool_call_id)]}
                    )
                
                # Validate specific constraint requirements
                if constraint_type == "concentration" and "nutrient" not in constraint:
                    return Command(
                        update={"messages": [ToolMessage(f"Error: Concentration constraint {i} missing 'nutrient' field", tool_call_id=tool_call_id)]}
                    )
                elif constraint_type == "daily_total":
                    if "attribute" not in constraint:
                        return Command(
                            update={"messages": [ToolMessage(f"Error: Daily total constraint {i} missing 'attribute' field", tool_call_id=tool_call_id)]}
                        )
                    if "target" not in constraint:
                        return Command(
                            update={"messages": [ToolMessage(f"Error: Daily total constraint {i} missing 'target' field", tool_call_id=tool_call_id)]}
                        )
            
            # Get user_id from config and access store
            user_id = config["configurable"].get("user_id")
            if not user_id:
                return Command(
                    update={"messages": [ToolMessage("Error: User ID not found in configuration", tool_call_id=tool_call_id)]}
                )

            store = get_store()

            # Try user feedbase first
            namespace = ("feedbases", user_id, feed_base_name)
            feedbase_entry = await store.aget(namespace, "data")

            # If not found, try system feedbase
            if not feedbase_entry and feed_base_name.startswith("default_"):
                system_namespace = ("system_feedbases", feed_base_name)
                feedbase_entry = await store.aget(system_namespace, "data")

            if not feedbase_entry:
                return Command(
                    update={"messages": [ToolMessage(f"Error: Feedbase '{feed_base_name}' not found. Please create it first using add_feed tool or use the system 'default' feedbase.", tool_call_id=tool_call_id)]}
                )

            feedbase_data = feedbase_entry.value
            feed_database = feedbase_data.get("feeds", {})
            if not feed_database:
                return Command(
                    update={"messages": [ToolMessage(f"Error: Feedbase '{feed_base_name}' is empty. Please add feeds first using add_feed tool.", tool_call_id=tool_call_id)]}
                )
            
            # Check if selected feeds exist
            missing_feeds = [f for f in selected_feeds if f not in feed_database]
            if missing_feeds:
                return Command(
                    update={"messages": [ToolMessage(f"Error: The following feeds are not in the database: {', '.join(missing_feeds)}. Available feeds: {', '.join(feed_database.keys())}", tool_call_id=tool_call_id)]}
                )
            
            # Extract daily intake if provided in constraints
            daily_intake_kg = None
            for constraint in nutritional_constraints:
                if constraint.get("type") == "daily_total" and constraint.get("attribute") == "dmi":
                    daily_intake_kg = constraint.get("target")
                    break
            
            # Run optimization
            optimizer = create_optimizer()
            optimizer.set_feeds(feed_database)
            
            optimization_result = await asyncio.to_thread(
                optimizer.optimize,
                nutritional_constraints=nutritional_constraints,
                selected_feeds=selected_feeds,
                feed_constraints=feed_constraints or {},
                optimization_goal=optimization_goal
            )
            
            # If optimization succeeded and daily intake is provided, calculate daily amounts
            if optimization_result.get("status") == "success" and daily_intake_kg is not None:
                # Calculate daily amounts for each feed
                updated_formulation = {}
                for feed_name, feed_data in optimization_result["formulation"].items():
                    percentage_dm = feed_data["percentage_dm"]
                    
                    # Get feed dry matter percentage
                    feed_dm_percent = feed_database[feed_name]["dm_percent"]
                    
                    # Calculate daily amount as-fed
                    # Formula: (feed_percentage_dm / 100) * daily_dm_intake_kg / (feed_dm_percent / 100)
                    kg_per_day = (percentage_dm / 100) * daily_intake_kg / (feed_dm_percent / 100)
                    
                    updated_formulation[feed_name] = {
                        "percentage_dm": percentage_dm,
                        "kg_per_day": round(kg_per_day, 2)
                    }
                
                # Update the formulation result
                optimization_result["formulation"] = updated_formulation
                optimization_result["daily_dm_intake_kg"] = daily_intake_kg
            
            # Only store constraints in state if optimization succeeded
            state_update = {
                "current_formulation": optimization_result,
                "current_feedbase_name": feed_base_name,  # Store feedbase reference for export
                "current_user_id": user_id,  # Store user ID for export
                "messages": [
                    ToolMessage(json.dumps(optimization_result, indent=2), tool_call_id=tool_call_id)
                ]
            }
            
            if optimization_result.get("status") == "success":
                state_update["formulation_constraints"] = nutritional_constraints
                state_update["feed_constraints"] = feed_constraints or {}
            
            return Command(update=state_update)
            
        except Exception as e:
            logger.error(f"Formulation error: {e}")
            return Command(
                update={"messages": [ToolMessage(f"Error formulating ration: {str(e)}", tool_call_id=tool_call_id)]}
            )
    
    @tool
    async def check_feeds(
        feed_base_name: str,
        state: Annotated[dict, InjectedState],
        tool_call_id: Annotated[str, InjectedToolCallId],
        config: RunnableConfig,
        query: str = ""
    ) -> Command:
        """
        Query feeds in a feedbase using SQL-like syntax or list syntax.
        
        Query Syntax Options:
        
        1. Empty query ("") - Returns category summary for large feedbases, or full list for small ones
        
        2. Special queries:
           - "nutrients" → Returns list of all available nutrient column names
        
        3. SQL-like query:
           [regex_pattern] WHERE category IN [cat1, cat2] ORDER BY nutrient ASC|DESC LIMIT n RETURN full
           
           - regex_pattern: Matches feed names (case-insensitive), e.g., "corn.*", "soy.*"
           - WHERE category IN [...]: Filter by category, e.g., WHERE category IN [Energy Source, Plant Protein]
           - ORDER BY nutrient ASC|DESC: Sort by nutrient value
           - LIMIT n: Max results (default 20, max 20)
           - RETURN full: Include all nutrients (default returns names only)
        
        4. List syntax:
           "[feed1, feed2, feed3]" → Returns full details for specific feeds
        
        Examples:
            check_feeds("default_dairy_cow", "")  
            # → Category summary (Energy Source: 30, Plant Protein: 21, ...)
            
            check_feeds("default_dairy_cow", "nutrients")  
            # → [Fd_CP, Fd_NDF, Fd_ADF, Fd_DE_Base, ...]
            
            check_feeds("default_dairy_cow", "corn.* WHERE category IN [Energy Source] LIMIT 10")
            # → corn_grain, corn_grain_hm, corn_silage, ... (names only)
            
            check_feeds("default_dairy_cow", "soy.* ORDER BY Fd_CP DESC LIMIT 5 RETURN full")
            # → Full nutrient data for matching feeds
            
            check_feeds("default_dairy_cow", "[corn_silage, alfalfa_hay, soybean_meal_48]")
            # → Full details for these specific feeds
        
        Args:
            feed_base_name: Name of the feedbase to query
            query: Query string (empty for summary, or SQL-like/list syntax)
            
        Returns:
            Query results (names by default, full nutrients with RETURN full or list syntax)
        """
        from services.session_manager import check_token_limit
        
        HUGE_FEEDBASE_THRESHOLD = 50
        MAX_LIMIT = 20
        MAX_TOKENS = 50000  # Safe limit for tool responses
        
        def parse_query(query_str: str) -> dict:
            """Parse SQL-like query into components."""
            result = {
                "pattern": None,
                "categories": [],
                "order_by": None,
                "order_dir": "ASC",
                "limit": MAX_LIMIT,
                "return_full": False  # Default to names only
            }
            
            if not query_str:
                return result
            
            q = query_str.strip()
            
            # Extract RETURN full clause
            return_match = re.search(r'\bRETURN\s+full\b', q, re.IGNORECASE)
            if return_match:
                result["return_full"] = True
                q = q[:return_match.start()] + q[return_match.end():]
            
            # Extract LIMIT clause
            limit_match = re.search(r'\bLIMIT\s+(\d+)\b', q, re.IGNORECASE)
            if limit_match:
                result["limit"] = min(int(limit_match.group(1)), MAX_LIMIT)
                q = q[:limit_match.start()] + q[limit_match.end():]
            
            # Extract ORDER BY clause
            order_match = re.search(r'\bORDER\s+BY\s+(\w+)\s*(ASC|DESC)?\b', q, re.IGNORECASE)
            if order_match:
                result["order_by"] = order_match.group(1)
                result["order_dir"] = (order_match.group(2) or "ASC").upper()
                q = q[:order_match.start()] + q[order_match.end():]
            
            # Extract WHERE category IN [...] clause
            where_match = re.search(r'\bWHERE\s+category\s+IN\s*\[([^\]]+)\]', q, re.IGNORECASE)
            if where_match:
                cats = where_match.group(1)
                result["categories"] = [c.strip() for c in cats.split(",") if c.strip()]
                q = q[:where_match.start()] + q[where_match.end():]
            
            # Remaining text is the pattern
            pattern = q.strip()
            if pattern:
                result["pattern"] = pattern
            
            return result
        
        def format_feed_names(feeds: list) -> str:
            """Format feed names as comma-separated list."""
            return ", ".join(feeds)
        
        def format_feed_full(feed_name: str, feed_data: dict) -> str:
            """Format feed with all nutrients (full mode), skipping 0/null values."""
            lines = [f"**{feed_name}**"]
            lines.append(f"  DM: {feed_data.get('dm_percent', 'N/A')}%")
            lines.append(f"  Cost: ${feed_data.get('cost_per_kg', 0)}/kg")
            
            if feed_data.get('category'):
                lines.append(f"  Category: {feed_data['category']}")
            if feed_data.get('type'):
                lines.append(f"  Type: {feed_data['type']}")
            
            nutrients = feed_data.get('nutrients', {})
            if nutrients:
                lines.append("  Nutrients (DM basis):")
                for nutrient, value in nutrients.items():
                    # Skip 0/null values
                    if value not in (0, 0.0, None, ""):
                        lines.append(f"    {nutrient}: {value}")
            return "\n".join(lines)
        
        try:
            # Validate feedbase name
            if not isinstance(feed_base_name, str) or not feed_base_name.strip():
                return Command(
                    update={"messages": [ToolMessage("Error: Feedbase name must be a non-empty string", tool_call_id=tool_call_id)]}
                )
            
            if _is_free_tier(config) and not feed_base_name.startswith("default_"):
                return _free_tier_feedbase_message("inspect non-system", tool_call_id)
            
            # Get user_id from config and access store
            user_id = config["configurable"].get("user_id")
            if not user_id:
                return Command(
                    update={"messages": [ToolMessage("Error: User ID not found in configuration", tool_call_id=tool_call_id)]}
                )

            store = get_store()

            # Try user feedbase first
            namespace = ("feedbases", user_id, feed_base_name)
            feedbase_entry = await store.aget(namespace, "data")

            # If not found, try system feedbase
            is_system_feedbase = False
            if not feedbase_entry and feed_base_name.startswith("default_"):
                system_namespace = ("system_feedbases", feed_base_name)
                feedbase_entry = await store.aget(system_namespace, "data")
                is_system_feedbase = True

            if not feedbase_entry:
                return Command(
                    update={"messages": [ToolMessage(f"Feedbase '{feed_base_name}' not found.", tool_call_id=tool_call_id)]}
                )

            feedbase_data = feedbase_entry.value
            feed_database = feedbase_data.get("feeds", {})
            feedbase_animal_type = feedbase_data.get("animal_type", "dairy_cow")

            if not feed_database:
                return Command(
                    update={"messages": [ToolMessage(f"Feedbase '{feed_base_name}' is empty.", tool_call_id=tool_call_id)]}
                )
            
            num_feeds = len(feed_database)
            is_huge = num_feeds > HUGE_FEEDBASE_THRESHOLD
            feedbase_type = "[SYSTEM]" if is_system_feedbase else "[USER]"
            header = f"Feedbase '{feed_base_name}' {feedbase_type} ({feedbase_animal_type}, {num_feeds} feeds)"
            
            query = query.strip() if query else ""
            
            # Handle special query: "nutrients"
            if query.lower() == "nutrients":
                # Collect all unique nutrient names from all feeds
                all_nutrients = set()
                for feed_data in feed_database.values():
                    nutrients = feed_data.get("nutrients", {})
                    all_nutrients.update(nutrients.keys())
                
                sorted_nutrients = sorted(all_nutrients)
                result = f"{header}\n\nAvailable nutrient columns ({len(sorted_nutrients)}):\n{', '.join(sorted_nutrients)}"
                return Command(
                    update={"messages": [ToolMessage(result, tool_call_id=tool_call_id)]}
                )
            
            # Handle list syntax: "[feed1, feed2, feed3]"
            list_match = re.match(r'^\s*\[([^\]]+)\]\s*$', query)
            if list_match:
                feed_names = [f.strip() for f in list_match.group(1).split(",") if f.strip()]
                output_lines = [header, ""]
                found_feeds = []
                missing_feeds = []
                
                for name in feed_names:
                    if name in feed_database:
                        found_feeds.append(name)
                    else:
                        missing_feeds.append(name)
                
                if missing_feeds:
                    output_lines.append(f"⚠️ Not found: {', '.join(missing_feeds)}\n")
                
                for name in found_feeds:
                    output_lines.append(format_feed_full(name, feed_database[name]))
                    output_lines.append("")
                
                result = "\n".join(output_lines)
                
                # Token check
                is_within_limit, token_count, error_msg = check_token_limit(result, MAX_TOKENS)
                if not is_within_limit:
                    return Command(
                        update={"messages": [ToolMessage(f"Error: {error_msg}\nTry requesting fewer feeds.", tool_call_id=tool_call_id)]}
                    )
                
                result = f"[FULL] {result}"
                return Command(
                    update={"messages": [ToolMessage(result, tool_call_id=tool_call_id)]}
                )
            
            # Handle empty query - return summary for huge feedbases
            if not query:
                if is_huge:
                    # Return category summary
                    categories = {}
                    for feed_name, feed_data in feed_database.items():
                        cat = feed_data.get("category", "Unknown")
                        categories[cat] = categories.get(cat, 0) + 1
                    
                    output_lines = [header, "", "Categories:"]
                    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
                        output_lines.append(f"  - {cat}: {count} feeds")
                    
                    output_lines.append("")
                    output_lines.append("ℹ️ This is a large feedbase. Use query syntax:")
                    output_lines.append(f'  check_feeds("{feed_base_name}", "nutrients")  # List nutrient columns')
                    output_lines.append(f'  check_feeds("{feed_base_name}", "corn.* LIMIT 10")  # Search by name')
                    output_lines.append(f'  check_feeds("{feed_base_name}", "WHERE category IN [Plant Protein]")')
                    output_lines.append(f'  check_feeds("{feed_base_name}", "[feed1, feed2]")  # Full details for specific feeds')
                    
                    result = "\n".join(output_lines)
                    return Command(
                        update={"messages": [ToolMessage(result, tool_call_id=tool_call_id)]}
                    )
                else:
                    # Small feedbase - return all feed names
                    feed_names = list(feed_database.keys())
                    output_lines = [header, "", format_feed_names(feed_names)]
                    result = "\n".join(output_lines)
                    
                    return Command(
                        update={"messages": [ToolMessage(result, tool_call_id=tool_call_id)]}
                    )
            
            # Parse SQL-like query
            parsed = parse_query(query)
            
            # Filter feeds
            filtered_feeds = []
            for feed_name, feed_data in feed_database.items():
                # Apply pattern filter
                if parsed["pattern"]:
                    try:
                        if not re.search(parsed["pattern"], feed_name, re.IGNORECASE):
                            # Also check nasem_name if available
                            nasem_name = feed_data.get("nasem_name", "")
                            if not re.search(parsed["pattern"], nasem_name, re.IGNORECASE):
                                continue
                    except re.error:
                        # If invalid regex, treat as substring match
                        if parsed["pattern"].lower() not in feed_name.lower():
                            nasem_name = feed_data.get("nasem_name", "").lower()
                            if parsed["pattern"].lower() not in nasem_name:
                                continue
                
                # Apply category filter
                if parsed["categories"]:
                    feed_cat = feed_data.get("category", "")
                    if not any(cat.lower() == feed_cat.lower() for cat in parsed["categories"]):
                        continue
                
                filtered_feeds.append((feed_name, feed_data))
            
            # Sort if requested
            if parsed["order_by"]:
                def get_sort_key(item):
                    _, feed_data = item
                    nutrients = feed_data.get("nutrients", {})
                    val = nutrients.get(parsed["order_by"], 0)
                    return val if val is not None else 0
                
                reverse = parsed["order_dir"] == "DESC"
                filtered_feeds.sort(key=get_sort_key, reverse=reverse)
            
            # Apply limit
            filtered_feeds = filtered_feeds[:parsed["limit"]]
            
            # Format output based on return type
            output_lines = [header, f"\nQuery: {query}", f"Results: {len(filtered_feeds)} feeds\n"]
            
            if parsed["return_full"]:
                # Full nutrient data
                for feed_name, feed_data in filtered_feeds:
                    output_lines.append(format_feed_full(feed_name, feed_data))
                    output_lines.append("")
                result = "\n".join(output_lines)
            else:
                # Names only (default)
                feed_names = [name for name, _ in filtered_feeds]
                output_lines.append(format_feed_names(feed_names))
                result = "\n".join(output_lines)
            
            # Token check for all responses
            is_within_limit, token_count, error_msg = check_token_limit(result, MAX_TOKENS)
            if not is_within_limit:
                return Command(
                    update={"messages": [ToolMessage(
                        f"Error: {error_msg}\n"
                        f"Try: LIMIT {parsed['limit'] // 2}",
                        tool_call_id=tool_call_id
                    )]}
                )
            
            return Command(
                update={"messages": [ToolMessage(result, tool_call_id=tool_call_id)]}
            )
            
        except Exception as e:
            logger.error(f"Check feeds error: {e}")
            return Command(
                update={"messages": [ToolMessage(f"Error retrieving feed information: {str(e)}", tool_call_id=tool_call_id)]}
            )

    
    
    @tool
    async def export_formulation(
        description: str,
        state: Annotated[dict, InjectedState],
        tool_call_id: Annotated[str, InjectedToolCallId],
        config: RunnableConfig,
        filename: Optional[str] = None,
        animal_input: Optional[Dict[str, Any]] = None
    ) -> Command:
        """
        Export current formulation to Excel with multi-tab layout.

        Creates Excel with 3 sheets (4 for dairy cows):
        1. Formulation Results
        2. Formulation Notes
        3. Constraints Used
        4. NASEM Analysis (dairy_cow only, if animal_input provided)

        Args:
            description: Detailed formulation description and recommendations (supports multi-line text)
            filename: Optional custom filename (default: formulation_export_TIMESTAMP.xlsx)
            animal_input: Optional animal parameters for NASEM evaluation (dairy_cow only).
                         Required keys: body_weight_kg, days_in_milk, parity, target_milk_kg
                         Optional: milk_fat_percent, milk_protein_percent, days_pregnant, breed

        Returns:
            Excel file with feeding information
        """
        try:
            # Get preferred language and texts
            preferred_language = config.get("configurable", {}).get("preferred_language", "zh-CN")
            locale = normalize_locale(preferred_language)
            texts = EXPORT_TRANSLATIONS.get(locale, EXPORT_TRANSLATIONS["zh-CN"])

            # Get all required data from state
            current_formulation = state.get("current_formulation", {})
            formulation_constraints = state.get("formulation_constraints", [])
            feed_constraints = state.get("feed_constraints", {})

            # Get feedbase references from state
            feedbase_name = state.get("current_feedbase_name", "")
            user_id = state.get("current_user_id", "")

            if not current_formulation or current_formulation.get("status") != "success":
                return Command(
                    update={"messages": [ToolMessage(texts["error_no_formulation"], tool_call_id=tool_call_id)]}
                )

            # Get feed database from store using state references
            feed_database = {}
            if feedbase_name and user_id:
                try:
                    store = get_store()

                    # Try user feedbase first
                    namespace = ("feedbases", user_id, feedbase_name)
                    feedbase_entry = await store.aget(namespace, "data")

                    # If not found, try system feedbase
                    if not feedbase_entry and feedbase_name.startswith("default_"):
                        system_namespace = ("system_feedbases", feedbase_name)
                        feedbase_entry = await store.aget(system_namespace, "data")

                    if feedbase_entry:
                        feedbase_data = feedbase_entry.value
                        feed_database = feedbase_data.get("feeds", {})
                except Exception as e:
                    logger.error(f"Failed to get feedbase data from store: {e}")

            if not feed_database:
                return Command(
                    update={"messages": [ToolMessage(texts["error_no_db"], tool_call_id=tool_call_id)]}
                )

            # Get session workspace path from config
            try:
                session_id = config["configurable"].get("thread_id")
                if session_id:
                    from services.session_manager import session_manager
                    workspace_path = await session_manager.get_session_workspace_path(session_id)
                else:
                    logger.warning("No session_id in config, using current directory")
                    workspace_path = Path(".")
            except Exception as e:
                logger.error(f"Failed to get session workspace: {e}")
                workspace_path = Path(".")

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"formulation_export_{timestamp}.xlsx"

            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = workspace_path / filename

            # Capture daily intake information if available
            daily_intake_kg = current_formulation.get("daily_dm_intake_kg")
            nutrient_analysis = current_formulation.get("nutrient_analysis", {})
            formulation_feeds = current_formulation.get("formulation", {})

            # Helper function to validate constraints and calculate achieved values
            def validate_constraints():
                constraint_results = []

                for constraint in formulation_constraints:
                    constraint_type = constraint.get("type", "")
                    result = {
                        "type": "",
                        "nutrient": "",
                        "condition": "",
                        "actual": "",
                        "unit": "",
                        "satisfaction": ""
                    }
                    
                    if constraint_type == "concentration":
                        nutrient = constraint["nutrient"]
                        min_val = constraint.get("min")
                        max_val = constraint.get("max")
                        achieved = nutrient_analysis.get(nutrient, 0.0)
                        
                        result["type"] = texts["con_concentration"]
                        result["nutrient"] = nutrient
                        result["actual"] = achieved
                        result["unit"] = "% DM"
                        
                        if min_val is not None and max_val is not None:
                            result["condition"] = f"{min_val} - {max_val}"
                            satisfied = min_val <= achieved <= max_val
                        elif min_val is not None:
                            result["condition"] = f"≥ {min_val}"
                            satisfied = achieved >= min_val
                        elif max_val is not None:
                            result["condition"] = f"≤ {max_val}"
                            satisfied = achieved <= max_val
                        else:
                            satisfied = True
                            
                        result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    elif constraint_type == "daily_total":
                        attribute = constraint.get("attribute")
                        target = constraint.get("target")
                        tolerance_percent = constraint.get("tolerance_percent", 10.0)
                        
                        if attribute == "dmi":
                            result["type"] = texts["con_dmi"]
                            result["nutrient"] = "DMI"
                            result["unit"] = "kg/day"
                            result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}%"
                            result["actual"] = f"{target} ({texts['range']})"
                            result["satisfaction"] = texts["satisfied"]
                        else:
                            # Nutrient daily total constraint
                            result["type"] = texts["con_daily"]
                            result["nutrient"] = attribute
                            result["unit"] = texts["con_daily"] # Daily Intake
                            
                            if daily_intake_kg:
                                nutrient_percent = nutrient_analysis.get(attribute, 0.0)
                                achieved = (nutrient_percent / 100) * daily_intake_kg
                                result["actual"] = round(achieved, 2)
                                
                                tolerance_factor = tolerance_percent / 100.0
                                target_min = target * (1 - tolerance_factor)
                                target_max = target * (1 + tolerance_factor)
                                result["condition"] = f"{texts['target']}: {target} ± {tolerance_percent}% ({target_min:.2f} - {target_max:.2f})"
                                satisfied = target_min <= achieved <= target_max
                            else:
                                result["actual"] = texts["daily_intake_needed"]
                                result["condition"] = texts["cannot_calculate"]
                                satisfied = False
                                
                            result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    elif constraint_type == "ratio":
                        numerator = constraint["numerator"]
                        denominator = constraint["denominator"]
                        min_ratio = constraint.get("min")
                        max_ratio = constraint.get("max")
                        
                        result["type"] = texts["con_ratio"]
                        result["nutrient"] = f"{numerator}:{denominator}"
                        result["unit"] = texts["con_ratio"] # Ratio
                        
                        num_content = nutrient_analysis.get(numerator, 0.0)
                        denom_content = nutrient_analysis.get(denominator, 0.0)
                        
                        if denom_content > 0:
                            achieved = num_content / denom_content
                            result["actual"] = round(achieved, 2)
                            
                            conditions = []
                            if min_ratio is not None:
                                conditions.append(f"≥ {min_ratio}")
                            if max_ratio is not None:
                                conditions.append(f"≤ {max_ratio}")
                            result["condition"] = " & ".join(conditions)
                            
                            satisfied = True
                            if min_ratio is not None and achieved < min_ratio:
                                satisfied = False
                            if max_ratio is not None and achieved > max_ratio:
                                satisfied = False
                        else:
                            result["actual"] = texts["denom_zero"]
                            result["condition"] = texts["cannot_calculate"]
                            satisfied = False
                        
                        result["satisfaction"] = texts["satisfied"] if satisfied else texts["unsatisfied"]
                    
                    constraint_results.append(result)
                
                return constraint_results
            
            # Get constraint validation results
            constraint_results = validate_constraints()

            # ==================== Run NASEM Evaluation First (dairy_cow only) ====================
            model_output = None
            predicted_milk = 0.0
            if animal_type == "dairy_cow" and animal_input:
                try:
                    from services.nasem_service import get_nasem_service
                    
                    nasem_service = get_nasem_service()
                    
                    # Build diet composition from formulation
                    diet_composition = {}
                    for feed_name, feed_data in formulation_feeds.items():
                        kg_per_day = feed_data.get("kg_per_day", 0)
                        if kg_per_day and kg_per_day > 0:
                            diet_composition[feed_name] = kg_per_day
                    
                    if diet_composition and feed_database:
                        # Build animal input for NASEM
                        nasem_animal_input = nasem_service.build_animal_input(
                            body_weight_kg=animal_input.get("body_weight_kg", 650),
                            days_in_milk=animal_input.get("days_in_milk", 100),
                            parity=animal_input.get("parity", 2),
                            target_milk_kg=animal_input.get("target_milk_kg", 35),
                            milk_fat_percent=animal_input.get("milk_fat_percent", 3.5),
                            milk_protein_percent=animal_input.get("milk_protein_percent", 3.2),
                            days_pregnant=animal_input.get("days_pregnant", 0),
                            breed=animal_input.get("breed", "Holstein")
                        )
                        
                        # Build feedbase dict for NASEM
                        feedbase_dict = {"feeds": feed_database}
                        
                        # Run NASEM evaluation with full output
                        nasem_results = nasem_service.evaluate_diet(
                            feedbase=feedbase_dict,
                            diet_composition=diet_composition,
                            animal_input=nasem_animal_input,
                            return_full_output=True
                        )
                        
                        if nasem_results.get("status") == "success":
                            model_output = nasem_results.get("model_output")
                            # Extract predicted milk for profitability
                            if model_output:
                                predicted_milk = model_output.get_value("Mlk_Prod_comp") or 0.0
                        
                except Exception as e:
                    logger.warning(f"NASEM evaluation failed: {e}")
                    model_output = None
                    predicted_milk = 0.0

            # Create Excel with tabs
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

                # ==================== SINGLE TAB: Summary (All sections combined) ====================
                # Layout: A-E: ingredients, nutrients, constraints | F-G: profitability | H+: notes
                # Constraints and NASEM summary are side-by-side at the bottom
                
                # Build main column data (A-E)
                main_rows = []
                
                # Header
                main_rows.append([texts["sheet_summary"], "", "", "", ""])
                main_rows.append([f"{texts['date']}:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", ""])
                main_rows.append(["", "", "", "", ""])
                
                # Section 1: Feed Formula Table with Costs
                main_rows.append([texts["ingredients"], "", "", "", ""])
                
                # Calculate total cost and prepare cost data
                total_cost = 0.0
                feed_rows = []
                for feed_name, feed_data in formulation_feeds.items():
                    sanitized_name = sanitize_feed_name(feed_name)
                    kg_per_day = feed_data.get("kg_per_day", 0)
                    percentage_dm = feed_data.get("percentage_dm", 0)
                    
                    # Get price from feed database
                    price_per_kg = 0.0
                    if feed_name in feed_database:
                        price_per_kg = feed_database[feed_name].get("price", 0) or 0
                    
                    cost_per_day = kg_per_day * price_per_kg if kg_per_day and price_per_kg else 0
                    total_cost += cost_per_day
                    
                    feed_rows.append([
                        sanitized_name,
                        round(kg_per_day, 2) if kg_per_day else "N/A",
                        round(percentage_dm, 1) if percentage_dm else "N/A",
                        round(price_per_kg, 2) if price_per_kg else "-",
                        round(cost_per_day, 2) if cost_per_day else "-"
                    ])
                
                # Header row for feed table
                main_rows.append([texts["feed_name"], texts["amount"], texts["dm_percent"], 
                                 texts["price_per_kg"], texts["cost_per_day"]])
                main_rows.extend(feed_rows)
                
                # Total row
                main_rows.append(["TOTAL", daily_intake_kg or "-", "100%", "", round(total_cost, 2)])
                main_rows.append(["", "", "", "", ""])
                
                # Section 2: Key Nutrients
                main_rows.append([texts["key_nutrients"], "", "", "", ""])
                
                # Key nutrients - map display names to NASEM field names
                key_nutrient_map = {
                    "CP": "Fd_CP",
                    "NDF": "Fd_NDF",
                    "ADF": "Fd_ADF",
                    "Fat": "Fd_CFat",
                    "Ca": "Fd_Ca",
                    "P": "Fd_P"
                }
                for display_name, field_name in key_nutrient_map.items():
                    value = nutrient_analysis.get(field_name, "-")
                    if isinstance(value, (int, float)):
                        value = round(value, 2)
                    main_rows.append([display_name, f"{value}%", "", "", ""])
                
                # F:C ratio
                forage_dm = 0.0
                concentrate_dm = 0.0
                for feed_name, feed_data in formulation_feeds.items():
                    kg = feed_data.get("kg_per_day", 0) or 0
                    if feed_name in feed_database:
                        category = feed_database[feed_name].get("type", "").lower()
                        if "forage" in category or "silage" in category or "hay" in category:
                            forage_dm += kg
                        else:
                            concentrate_dm += kg
                
                if forage_dm + concentrate_dm > 0:
                    forage_pct = round((forage_dm / (forage_dm + concentrate_dm)) * 100)
                    concentrate_pct = 100 - forage_pct
                    main_rows.append([texts["fc_ratio"], f"{forage_pct}:{concentrate_pct}", "", "", ""])
                
                main_rows.append(["", "", "", "", ""])
                
                # Track where constraints start for side-by-side with NASEM
                constraints_start_row = len(main_rows)
                
                # Section 3: Constraints (in main column area, will be side by side with NASEM)
                main_rows.append([texts["sheet_constraints"], "", "", "", ""])
                main_rows.append([texts["constraint_type"], texts["nutrient"], texts["condition"], texts["actual"], texts["satisfaction"]])
                
                for result in constraint_results:
                    main_rows.append([
                        result["type"],
                        result["nutrient"],
                        result["condition"],
                        result["actual"],
                        result["satisfaction"]
                    ])
                
                main_rows.append(["", "", "", "", ""])
                
                # Feed constraints
                main_rows.append([texts["feed_constraints"], "", "", "", ""])
                if feed_constraints:
                    main_rows.append([texts["feed_name"], texts["min_percent"], texts["max_percent"], "", ""])
                    for feed_name, constraints in feed_constraints.items():
                        sanitized_name = sanitize_feed_name(feed_name)
                        min_val = constraints.get('min', '')
                        max_val = constraints.get('max', '')
                        main_rows.append([sanitized_name, min_val, max_val, "", ""])
                else:
                    main_rows.append([texts["no_feed_constraints"], "", "", "", ""])
                
                # Build profitability column data (G-H, starting at row 1)
                profit_rows = []
                profit_rows.append([texts["profitability"], ""])
                profit_rows.append([texts["input_section"], ""])
                profit_rows.append([texts["herd_size"], 100])  # Default 100
                profit_rows.append([texts["milk_price"], 4.0])  # Default 4.0 yuan/kg
                profit_rows.append(["", ""])
                
                # Cost metrics
                cost_per_kg_dm = round(total_cost / daily_intake_kg, 2) if daily_intake_kg and daily_intake_kg > 0 else 0
                profit_rows.append([texts["cost_per_kg_dm"], cost_per_kg_dm])
                profit_rows.append([texts["cost_per_cow_day"], round(total_cost, 2)])
                profit_rows.append([texts["nasem_predicted_milk"], round(predicted_milk, 2) if predicted_milk else 0])
                profit_rows.append([texts["revenue_per_cow_day"], 0])  # Formula placeholder
                profit_rows.append([texts["profit_per_cow_day"], 0])  # Formula placeholder
                profit_rows.append(["", ""])
                profit_rows.append([texts["herd_profit_day"], 0])  # Formula placeholder
                profit_rows.append([texts["herd_profit_month"], 0])  # Formula placeholder
                
                # Build notes as single text (will be put in vertically merged cell F)
                # Join all lines with newline for the merged cell
                if isinstance(description, str) and description.strip():
                    notes_text = description
                else:
                    notes_text = texts["notes_none"]
                
                # Build NASEM summary for side column (starting at constraints_start_row in column G-H)
                nasem_summary_rows = []
                if model_output is not None:
                    nasem_summary_rows.append([texts["nasem_production"], ""])
                    
                    # Get values using ModelOutput.get_value()
                    mlk_prod = model_output.get_value("Mlk_Prod_comp")
                    mp_allow = model_output.get_value("Mlk_Prod_MPalow")
                    ne_allow = model_output.get_value("Mlk_Prod_NEalow")
                    milk_fat = model_output.get_value("MlkFat_Milk")
                    milk_protein = model_output.get_value("MlkNP_Milk")
                    
                    # Limiting factor
                    try:
                        mp_val = float(mp_allow) if mp_allow else 999
                        ne_val = float(ne_allow) if ne_allow else 999
                        if mp_val < ne_val:
                            limiting_factor = "MP (protein)"
                        elif ne_val < mp_val:
                            limiting_factor = "NE (energy)"
                        else:
                            limiting_factor = "Balanced"
                    except:
                        limiting_factor = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_predicted_milk"], round(mlk_prod, 2) if mlk_prod else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_allow_milk"], round(mp_allow, 2) if mp_allow else "N/A"])
                    nasem_summary_rows.append([texts["nasem_ne_allow_milk"], round(ne_allow, 2) if ne_allow else "N/A"])
                    nasem_summary_rows.append([texts["nasem_limiting_factor"], limiting_factor])
                    nasem_summary_rows.append([texts["nasem_milk_fat"], round(milk_fat, 4) if milk_fat else "N/A"])
                    nasem_summary_rows.append([texts["nasem_milk_protein"], round(milk_protein, 4) if milk_protein else "N/A"])
                    nasem_summary_rows.append(["", ""])
                    
                    # Energy Balance
                    nasem_summary_rows.append([texts["nasem_energy_balance"], ""])
                    me_intake = model_output.get_value("An_MEIn")
                    me_required = model_output.get_value("Trg_MEuse")
                    try:
                        me_balance = round(float(me_intake) - float(me_required), 2)
                        me_balance_str = f"+{me_balance}" if me_balance >= 0 else str(me_balance)
                    except:
                        me_balance_str = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_me_intake"], round(me_intake, 2) if me_intake else "N/A"])
                    nasem_summary_rows.append([texts["nasem_me_required"], round(me_required, 2) if me_required else "N/A"])
                    nasem_summary_rows.append([texts["nasem_me_balance"], me_balance_str])
                    nasem_summary_rows.append(["", ""])
                    
                    # Protein Balance
                    nasem_summary_rows.append([texts["nasem_protein_balance"], ""])
                    mp_intake = model_output.get_value("An_MPIn_g")
                    mp_required = model_output.get_value("An_MPuse_g_Trg")
                    rdp_intake = model_output.get_value("An_RDPIn_g")
                    try:
                        mp_balance = round(float(mp_intake) - float(mp_required), 1)
                        mp_balance_str = f"+{mp_balance}" if mp_balance >= 0 else str(mp_balance)
                    except:
                        mp_balance_str = "N/A"
                    
                    nasem_summary_rows.append([texts["nasem_mp_intake"], round(mp_intake, 1) if mp_intake else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_required"], round(mp_required, 1) if mp_required else "N/A"])
                    nasem_summary_rows.append([texts["nasem_mp_balance"], mp_balance_str])
                    nasem_summary_rows.append([texts["nasem_rdp_intake"], round(rdp_intake, 1) if rdp_intake else "N/A"])
                    nasem_summary_rows.append(["", ""])
                    
                    # Other Key Metrics
                    nasem_summary_rows.append([texts["nasem_other_metrics"], ""])
                    dcad = model_output.get_value("An_DCADmeq")
                    nasem_summary_rows.append([texts["nasem_dcad"], round(dcad, 1) if dcad else "N/A"])
                
                # Determine max rows needed
                max_rows = max(len(main_rows), len(profit_rows), constraints_start_row + len(nasem_summary_rows))
                
                # Track notes cell merge range for later formatting (starts at row 2 since row 1 is header)
                notes_merge_start_row = 2
                notes_merge_end_row = max_rows
                
                # Create combined data with proper positioning
                # Layout: A-E (main) | F (notes with header) | G-H (profitability/NASEM)
                # Borders will be added between sections in formatting phase
                combined_data = []
                for i in range(max_rows):
                    row = []
                    
                    # Columns A-E (main content)
                    if i < len(main_rows):
                        row.extend(main_rows[i])
                    else:
                        row.extend(["", "", "", "", ""])
                    
                    # Column F (notes - header in row 1, text in row 2 for merge)
                    if i == 0:
                        row.append(texts["sheet_notes"])  # Header
                    elif i == 1:
                        row.append(notes_text)  # Notes text starts at row 2
                    else:
                        row.append("")  # Empty for merge
                    
                    # Columns G-H (profitability at top, NASEM summary at constraints_start_row)
                    if i < len(profit_rows):
                        # Profitability section at top
                        row.extend(profit_rows[i])
                    elif i >= constraints_start_row and nasem_summary_rows:
                        # NASEM summary parallel to constraints
                        nasem_idx = i - constraints_start_row
                        if nasem_idx < len(nasem_summary_rows):
                            row.extend(nasem_summary_rows[nasem_idx])
                        else:
                            row.extend(["", ""])
                    else:
                        row.extend(["", ""])
                    
                    combined_data.append(row)
                
                # Write combined data
                combined_df = pd.DataFrame(combined_data)
                combined_df.to_excel(writer, sheet_name=texts["sheet_summary"], index=False, header=False)
                
                # ==================== NASEM Category Tabs (dairy_cow only) ====================
                # Note: model_output was obtained earlier before Excel creation
                
                # Create NASEM category tabs if we have full output
                if model_output is not None:
                    # Helper function to flatten nested dict to list of [key, value] rows
                    # Skips 0 values, None, and empty values for cleaner output
                    def flatten_dict(d, prefix="", skip_zeros=True):
                        rows = []
                        for k, v in d.items():
                            # Build display key (show nested path)
                            display_key = f"{prefix}.{k}" if prefix else k
                            
                            if isinstance(v, dict):
                                rows.extend(flatten_dict(v, display_key, skip_zeros))
                            elif isinstance(v, pd.DataFrame):
                                # Skip DataFrames (too complex for simple key-value)
                                continue
                            elif isinstance(v, (list, pd.Series)):
                                # Format arrays - skip if empty
                                if hasattr(v, 'tolist'):
                                    v = v.tolist()
                                if v and len(v) > 0 and len(str(v)) < 100:
                                    rows.append([display_key, str(v)])
                            elif v is not None:
                                # Skip zeros if requested
                                if skip_zeros and isinstance(v, (int, float)) and v == 0:
                                    continue
                                # Skip empty strings
                                if isinstance(v, str) and not v.strip():
                                    continue
                                # Round floats for cleaner display
                                if isinstance(v, float):
                                    v = round(v, 3)
                                rows.append([display_key, v])
                        return rows
                    


                    # NASEM category to sheet name mapping (NASEM-Summary is now in main tab)
                    nasem_categories = {
                        "Intakes": "NASEM-Intakes",
                        "Requirements": "NASEM-Requirements", 
                        "Production": "NASEM-Production",
                        "Excretion": "NASEM-Excretion",
                        "Efficiencies": "NASEM-Efficiencies"
                    }

                    
                    # Export each category as a separate tab
                    for category_name, sheet_name in nasem_categories.items():
                        try:
                            category_data = getattr(model_output, category_name, None)
                            if category_data is None:
                                continue
                            
                            tab_data = []
                            tab_data.append([sheet_name])
                            tab_data.append([f"{texts['date']}:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                            tab_data.append([])
                            tab_data.append(["Variable", "Value"])
                            
                            # Flatten the category dict
                            if isinstance(category_data, dict):
                                rows = flatten_dict(category_data)
                                tab_data.extend(rows)
                            
                            # Write tab
                            if len(tab_data) > 4:  # Only if we have data beyond headers
                                tab_df = pd.DataFrame(tab_data)
                                tab_df.to_excel(writer, sheet_name=sheet_name[:31], index=False, header=False)
                                
                        except Exception as e:
                            logger.warning(f"Failed to export NASEM {category_name}: {e}")
                            continue
                

                # Apply formatting
                workbook = writer.book
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import PieChart, Reference
                from openpyxl.utils import get_column_letter


                # Define common styles
                title_font = Font(bold=True, size=16, color="FFFFFF")
                title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                section_font = Font(bold=True, size=12, color="FFFFFF")
                section_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                label_font = Font(bold=True)
                thin_side = Side(style="thin", color="B4C6E7")
                table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                # ==================== FORMAT Summary Tab (unified) ====================
                if texts["sheet_summary"] in workbook.sheetnames:
                    ws = workbook[texts["sheet_summary"]]
                    
                    # Define yellow fill for editable input cells
                    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    # Define thick border for section dividers
                    thick_side = Side(style="medium", color="1F4E79")

                    # Auto-adjust column widths for new layout
                    # A-E: main content | F: notes | G-H: profitability/NASEM
                    ws.column_dimensions['A'].width = 30  # Main labels
                    ws.column_dimensions['B'].width = 15  # Main values
                    ws.column_dimensions['C'].width = 12  # Condition/DM%
                    ws.column_dimensions['D'].width = 12  # Actual/Price
                    ws.column_dimensions['E'].width = 15  # Satisfaction/Cost
                    ws.column_dimensions['F'].width = 50  # Notes column (wide)
                    ws.column_dimensions['G'].width = 28  # Profitability/NASEM labels
                    ws.column_dimensions['H'].width = 15  # Profitability/NASEM values

                    # Format notes header (F1) and merge notes text (F2:F{max_rows})
                    notes_header_cell = ws.cell(row=1, column=6)
                    notes_header_cell.font = section_font
                    notes_header_cell.fill = section_fill
                    
                    if notes_merge_end_row > notes_merge_start_row:
                        ws.merge_cells(f'F{notes_merge_start_row}:F{notes_merge_end_row}')
                        notes_cell = ws.cell(row=notes_merge_start_row, column=6)
                        notes_cell.font = Font(size=10)
                        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # Add outer borders around each section
                    max_row = ws.max_row
                    
                    # Helper to apply border preserving existing borders
                    def apply_border(cell, **sides):
                        existing = cell.border
                        cell.border = Border(
                            left=sides.get('left', existing.left),
                            right=sides.get('right', existing.right),
                            top=sides.get('top', existing.top),
                            bottom=sides.get('bottom', existing.bottom)
                        )
                    
                    # Apply borders to all cells
                    for row_num in range(1, max_row + 1):
                        for col_num in range(1, 9):  # A-H
                            cell = ws.cell(row=row_num, column=col_num)
                            
                            # Determine which borders to apply
                            left = thick_side if col_num == 1 else None  # Left edge of main
                            right = None
                            top = thick_side if row_num == 1 else None  # Top edge
                            bottom = thick_side if row_num == max_row else None  # Bottom edge
                            
                            # Right dividers between sections
                            if col_num == 5:  # E - divider after main
                                right = thick_side
                            elif col_num == 6:  # F - divider after notes
                                right = thick_side
                            elif col_num == 8:  # H - right edge of profitability
                                right = thick_side
                            
                            apply_border(cell, left=left, right=right, top=top, bottom=bottom)

                    # Track special cells for formula insertion in column H (profitability values)
                    herd_size_cell = None
                    milk_price_cell = None
                    predicted_milk_cell = None
                    cost_per_cow_cell = None
                    profit_row = None
                    
                    # Apply formatting based on content - scan all columns
                    for row_num in range(1, ws.max_row + 1):
                        # Column A content (main area)
                        a_cell = ws.cell(row=row_num, column=1)
                        a_value = str(a_cell.value or "")
                        b_cell = ws.cell(row=row_num, column=2)
                        
                        # Column G content (profitability/NASEM labels)
                        g_cell = ws.cell(row=row_num, column=7)
                        g_value = str(g_cell.value or "")
                        h_cell = ws.cell(row=row_num, column=8)

                        # ===== Main area formatting (column A) =====
                        # Title row
                        if a_value == texts["sheet_summary"]:
                            a_cell.font = title_font
                            a_cell.fill = title_fill
                            ws.merge_cells(f'A{row_num}:E{row_num}')

                        # Section headers in main area
                        elif a_value in [texts["ingredients"], texts["key_nutrients"], 
                                        texts["sheet_constraints"], texts["feed_constraints"]]:
                            a_cell.font = section_font
                            a_cell.fill = section_fill
                            ws.merge_cells(f'A{row_num}:E{row_num}')

                        # Feed table header
                        elif a_value == texts["feed_name"]:
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")
                        
                        # Constraint table header
                        elif a_value == texts["constraint_type"]:
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")

                        # TOTAL row
                        elif a_value == "TOTAL":
                            for col in range(1, 6):
                                cell = ws.cell(row=row_num, column=col)
                                cell.font = Font(bold=True)

                        # Label rows in main area (ending with :)
                        elif a_value.endswith(":"):
                            a_cell.font = label_font
                        
                        # Satisfaction pass/fail coloring in column E
                        e_cell = ws.cell(row=row_num, column=5)
                        e_value = str(e_cell.value or "")
                        if texts["satisfied"] in e_value:
                            e_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            e_cell.font = Font(color="006100")
                        elif texts["unsatisfied"] in e_value:
                            e_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            e_cell.font = Font(color="9C0006")

                        # ===== G-H area formatting (profitability + NASEM) =====
                        # Profitability/NASEM section headers (now in column G)
                        if g_value in [texts["profitability"], texts["nasem_production"], 
                                      texts["nasem_energy_balance"], texts["nasem_protein_balance"],
                                      texts["nasem_other_metrics"]]:
                            g_cell.font = section_font
                            g_cell.fill = section_fill
                            ws.merge_cells(f'G{row_num}:H{row_num}')

                        # Input cells - highlight yellow and track for formulas (values now in H)
                        elif g_value == texts["herd_size"]:
                            h_cell.fill = input_fill
                            herd_size_cell = f"H{row_num}"
                        elif g_value == texts["milk_price"]:
                            h_cell.fill = input_fill
                            milk_price_cell = f"H{row_num}"
                        
                        # Track cells for formulas
                        elif g_value == texts["nasem_predicted_milk"]:
                            predicted_milk_cell = f"H{row_num}"
                        elif g_value == texts["cost_per_cow_day"]:
                            cost_per_cow_cell = f"H{row_num}"
                        
                        # Revenue formula
                        elif g_value == texts["revenue_per_cow_day"]:
                            if predicted_milk_cell and milk_price_cell:
                                h_cell.value = f"={predicted_milk_cell}*{milk_price_cell}"
                        
                        # Profit per cow formula
                        elif g_value == texts["profit_per_cow_day"]:
                            if predicted_milk_cell and milk_price_cell and cost_per_cow_cell:
                                h_cell.value = f"={predicted_milk_cell}*{milk_price_cell}-{cost_per_cow_cell}"
                            profit_row = row_num
                        
                        # Herd profit/day formula
                        elif g_value == texts["herd_profit_day"]:
                            if profit_row and herd_size_cell:
                                h_cell.value = f"=H{profit_row}*{herd_size_cell}"
                        
                        # Herd profit/month formula
                        elif g_value == texts["herd_profit_month"]:
                            herd_day_cell = f"H{row_num - 1}" if row_num > 1 else None
                            if herd_day_cell:
                                h_cell.value = f"={herd_day_cell}*30"

                        # Input section label
                        elif g_value == texts["input_section"]:
                            g_cell.font = Font(italic=True, color="666666")

                        # Label rows in G column (ending with :)
                        elif g_value.endswith(":"):
                            g_cell.font = label_font

                # ==================== FORMAT NASEM Category Tabs ====================
                # Note: NASEM-Summary is now in main tab, so not included here
                nasem_sheet_names = ["NASEM-Intakes", "NASEM-Requirements", "NASEM-Production", 
                                     "NASEM-Excretion", "NASEM-Efficiencies"]
                
                for sheet_name in nasem_sheet_names:
                    if sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        
                        # Set column widths
                        ws.column_dimensions['A'].width = 50
                        ws.column_dimensions['B'].width = 25
                        
                        for row_num in range(1, ws.max_row + 1):
                            first_cell = ws.cell(row=row_num, column=1)
                            first_value = str(first_cell.value or "")
                            
                            # Title row (first row with sheet name)
                            if first_value == sheet_name:
                                first_cell.font = title_font
                                first_cell.fill = title_fill
                                ws.merge_cells(f'A{row_num}:B{row_num}')
                            
                            # Header row
                            elif first_value == "Variable":
                                first_cell.font = header_font
                                first_cell.fill = header_fill
                                second_cell = ws.cell(row=row_num, column=2)
                                if second_cell.value:
                                    second_cell.font = header_font
                                    second_cell.fill = header_fill
                            
                            # Label rows (ending with :)
                            elif first_value.endswith(":"):
                                first_cell.font = label_font
            


            # Format file info for backend parsing
            # Convert literal \n from LLM to actual newlines
            normalized_description = description.replace('\\n', '\n') if description else None

            file_info = {
                "filepath": str(filepath),
                "filename": filename,
                "type": "excel",
                "description": normalized_description
            }

            return Command(
                update={"messages": [ToolMessage(texts["export_success"].format(filename=filename) + f" [FILE_EXPORT]{json.dumps(file_info, ensure_ascii=False)}[/FILE_EXPORT]", tool_call_id=tool_call_id)]}
            )
            
        except Exception as e:
            logger.error(f"Export formulation error: {e}")
            return Command(
                update={"messages": [ToolMessage(texts["export_fail"].format(error=str(e)), tool_call_id=tool_call_id)]}
            )

    return [add_feed, formulate_ration, check_feeds, list_feed_bases, export_formulation]